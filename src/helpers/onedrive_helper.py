import os
from io import BytesIO

import msal
import pandas as pd
import requests
from dotenv import load_dotenv
from openpyxl import load_workbook

load_dotenv()

# Constants

CLIENT_ID = os.getenv("MS_CLIENT_ID")
CLIENT_SECRET = os.getenv("MS_CLIENT_SECRET")
TENANT_ID = os.getenv("MS_TENANT_ID")
USER_EMAIL = os.getenv("MS_USER_EMAIL")

AUTHORITY = f"https://login.microsoftonline.com/{TENANT_ID}"
SCOPES = ["https://graph.microsoft.com/.default"]

# Microsoft Graph API base URL for OneDrive file operations
GRAPH_DRIVE_URL = f"https://graph.microsoft.com/v1.0/users/{USER_EMAIL}/drive/root"

# Content-Type header value for Excel files
EXCEL_CONTENT_TYPE = (
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
)


# Authentication

def get_access_token() -> str:
    """
    Authenticate with Microsoft Identity Platform and return a Bearer access token.

    Uses client credentials flow via MSAL. Raises on authentication failure.
    """
    app = msal.ConfidentialClientApplication(
        client_id=CLIENT_ID,
        authority=AUTHORITY,
        client_credential=CLIENT_SECRET,
    )

    result = app.acquire_token_for_client(scopes=SCOPES)

    if "access_token" not in result:
        raise Exception(f"[Auth] Failed to obtain access token: {result}")

    return result["access_token"]


# File Download

def download_excel_from_onedrive(access_token: str, file_path: str) -> BytesIO | None:
    """
    Download an Excel file from OneDrive and return it as an in-memory buffer.

    Returns None if the file does not exist (HTTP 404), or raises on other errors.
    """
    headers = {"Authorization": f"Bearer {access_token}"}
    url = f"{GRAPH_DRIVE_URL}:{file_path}:/content"
    response = requests.get(url, headers=headers)

    if response.status_code == 200:
        print(f"[Download] File downloaded successfully: {file_path}")
        return BytesIO(response.content)

    if response.status_code == 404:
        print(f"[Download] File not found: {file_path}")
        return None

    raise Exception(f"[Download] Failed ({response.status_code}): {response.text}")


# File Upload

def upload_excel_to_onedrive(
    access_token: str,
    file_path: str,
    excel_buffer: BytesIO,
) -> bool:
    """
    Upload an in-memory Excel buffer to OneDrive, overwriting any existing file.

    Returns True on success, raises on failure.
    """
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Content-Type":  EXCEL_CONTENT_TYPE,
    }
    url      = f"{GRAPH_DRIVE_URL}:{file_path}:/content"
    response = requests.put(url, headers=headers, data=excel_buffer.getvalue())

    if response.status_code in [200, 201]:
        print(f"[Upload] File uploaded successfully: {file_path}")
        return True

    raise Exception(f"[Upload] Failed ({response.status_code}): {response.text}")


# Sheet Reading

def read_excel_sheet_from_onedrive(
    access_token: str,
    file_path: str,
    sheet_name: str,
) -> pd.DataFrame:
    """
    Download an Excel file from OneDrive and return a single sheet as a DataFrame.

    Returns an empty DataFrame if the file does not exist or the sheet cannot be read.
    """
    excel_buffer = download_excel_from_onedrive(access_token, file_path)

    if excel_buffer is None:
        return pd.DataFrame()

    try:
        df = pd.read_excel(excel_buffer, sheet_name=sheet_name)
        print(f"[Read] Sheet '{sheet_name}' read successfully — {len(df)} row(s).")
        return df
    except Exception as exc:
        print(f"[Read] Failed to read sheet '{sheet_name}': {exc}")
        return pd.DataFrame()


# Multi-Sheet Write

def write_multiple_sheets_to_onedrive(
    access_token: str,
    file_path: str,
    sheets_dict: dict[str, pd.DataFrame],
) -> None:
    """
    Write multiple DataFrames as sheets into an Excel file on OneDrive.

    If the file already exists, existing sheets are updated in-place. If the
    file does not exist or the update fails, a fresh file is created instead.
    """
    print(f"\n[Write] Preparing Excel file with {len(sheets_dict)} sheet(s)...")

    try:
        excel_buffer = download_excel_from_onedrive(access_token, file_path)

        if excel_buffer is not None:
            print("[Write] File exists — updating in place...")

            try:
                wb = load_workbook(excel_buffer)

                visible_sheets = [s for s in wb.worksheets if s.sheet_state == "visible"]

                if len(visible_sheets) == 0:
                    print("[Write] All sheets are hidden — unhiding the first sheet...")
                    if len(wb.worksheets) > 0:
                        wb.worksheets[0].sheet_state = "visible"
                        wb.active = 0

                for sheet in wb.worksheets:
                    if sheet.sheet_state != "visible":
                        sheet.sheet_state = "visible"

                temp_buffer = BytesIO()
                wb.save(temp_buffer)
                wb.close()
                temp_buffer.seek(0)

                output_buffer = BytesIO()

                with pd.ExcelWriter(temp_buffer, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                    for sheet_name, df in sheets_dict.items():
                        print(f"  [Write] Updating sheet: '{sheet_name}' ({len(df)} row(s))")
                        df.to_excel(writer, sheet_name=sheet_name, index=False)

                output_buffer = temp_buffer

            except Exception as exc:
                print(f"[Write] Error updating existing file: {exc}")
                print("[Write] Fallback — creating new file...")

                output_buffer = BytesIO()
                with pd.ExcelWriter(output_buffer, engine="openpyxl", mode="w") as writer:
                    for sheet_name, df in sheets_dict.items():
                        print(f"  [Write] Writing sheet: '{sheet_name}' ({len(df)} row(s))")
                        df.to_excel(writer, sheet_name=sheet_name, index=False)

        else:
            print("[Write] File does not exist — creating new file...")

            output_buffer = BytesIO()
            with pd.ExcelWriter(output_buffer, engine="openpyxl", mode="w") as writer:
                for sheet_name, df in sheets_dict.items():
                    print(f"  [Write] Writing sheet: '{sheet_name}' ({len(df)} row(s))")
                    df.to_excel(writer, sheet_name=sheet_name, index=False)

        output_buffer.seek(0)

        print(f"\n[Write] Uploading to OneDrive: {file_path}")
        print("[Write] Refreshing access token before upload...")
        fresh_token = get_access_token()
        upload_excel_to_onedrive(fresh_token, file_path, output_buffer)

        print("[Write] Upload complete.")

    except Exception as exc:
        print(f"[Write] Fatal error in write_multiple_sheets_to_onedrive: {exc}")
        raise


# Excel Writer Helper

def create_excel_writer_from_onedrive(
    access_token: str,
    file_path: str,
) -> tuple[BytesIO, str]:
    """
    Download an existing OneDrive Excel file and return a (buffer, mode) tuple.

    Returns ("w") mode for a new file, or ("a") mode for an existing one.
    Ensures hidden sheets are unhidden before returning the buffer for append mode.
    """
    excel_buffer = download_excel_from_onedrive(access_token, file_path)

    if excel_buffer is None:
        print("[Writer] File does not exist — new file will be created.")
        return BytesIO(), "w"

    print("[Writer] File exists — opening in append mode.")

    try:
        wb = load_workbook(excel_buffer)
        visible_sheets = [s for s in wb.worksheets if s.sheet_state == "visible"]

        if len(visible_sheets) == 0:
            print("[Writer] Fixing hidden sheets...")
            wb.worksheets[0].sheet_state = "visible"
            wb.active = 0

            temp_buffer = BytesIO()
            wb.save(temp_buffer)
            wb.close()
            temp_buffer.seek(0)

            return temp_buffer, "a"

        wb.close()

    except Exception as exc:
        print(f"[Writer] Warning while checking file: {exc}")

    return excel_buffer, "a"


# Hidden Sheet Repair

def fix_hidden_sheets_onedrive(access_token: str, file_path: str) -> bool:
    """
    Unhide all hidden worksheets in a OneDrive Excel file and re-upload it.

    Returns True on success, False if the file is not found or an error occurs.
    """
    print(f"[Fix] Repairing hidden sheets in: {file_path}")

    try:
        excel_buffer = download_excel_from_onedrive(access_token, file_path)

        if excel_buffer is None:
            print("[Fix] File not found.")
            return False

        wb = load_workbook(excel_buffer)
        print(f"[Fix] Total sheets: {len(wb.worksheets)}")

        fixed_count = 0
        for idx, sheet in enumerate(wb.worksheets):
            if sheet.sheet_state != "visible":
                sheet.sheet_state = "visible"
                fixed_count += 1
                print(f"  [Fix] Unhid sheet: '{sheet.title}'")

            if idx == 0:
                wb.active = 0

        print(f"[Fix] {fixed_count} sheet(s) unhidden.")

        output_buffer = BytesIO()
        wb.save(output_buffer)
        wb.close()
        output_buffer.seek(0)

        fresh_token = get_access_token()
        upload_excel_to_onedrive(fresh_token, file_path, output_buffer)

        print("[Fix] File repaired and uploaded successfully.")
        return True

    except Exception as exc:
        print(f"[Fix] Error: {exc}")
        return False


# Script Entry Point

if __name__ == "__main__":
    load_dotenv()
    token = get_access_token()
    print(f"[Main] Access token obtained: {token[:20]}...")