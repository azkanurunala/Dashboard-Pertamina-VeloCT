import os
import sys
import traceback
from datetime import datetime
from io import BytesIO

import pandas as pd
import requests
from dotenv import load_dotenv
from openpyxl import load_workbook

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from helpers.onedrive_helper import (
    download_excel_from_onedrive,
    get_access_token,
    upload_excel_to_onedrive,
)

load_dotenv()


# Constants

ONEDRIVE_FILE_PATH = os.getenv("ONEDRIVE_DATA_PATH", "/results/(Terstruktur)Data_Scraping_final.xlsx")

# REVISI 1: URL lama (sipsn.kemenlh.go.id) mati sejak Oktober 2024 akibat
# pemecahan KLHK via Perpres 139/2024. Ganti ke mirror legacy yang masih aktif.
API_URL = "https://sampahnasional.kemenlh.go.id/indikatif/public/home/ajax_list"

SHEET_MAPPING = {
    "sumber":    "(Data)WTE_Sumber",
    "komposisi": "(Data)WTE_Komposisi",
    "timbulan":  "(Data)WTE_Timbulan",
}

COLUMN_RENAME = {
    "nama_dati2":     "Nama Kota/Kabupaten",
    "nama_propinsi":  "Nama Provinsi",
}

DEDUP_ID_COLS = ["tahun", "Nama Provinsi", "Nama Kota/Kabupaten"]

# REVISI 2: Tambah headers agar request tidak diblok (domain baru lebih ketat)
HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "application/json, text/javascript, */*; q=0.01",
    "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
    "X-Requested-With": "XMLHttpRequest",
    "Referer": "https://sampahnasional.kemenlh.go.id/indikatif",
}


# Duplicate Check

def check_year_exists_in_onedrive(access_token, tahun: str) -> dict[str, bool]:
    """
    Check whether data for the given year already exists in each WTE sheet.

    Returns a dict {jenis: bool} for 'sumber', 'komposisi', 'timbulan'.
    """
    default = {jenis: False for jenis in SHEET_MAPPING}

    try:
        excel_buffer = download_excel_from_onedrive(access_token, ONEDRIVE_FILE_PATH)
        if excel_buffer is None:
            print("[Check] File belum ada di OneDrive.")
            return default

        excel_buffer.seek(0)
        wb      = load_workbook(excel_buffer)
        results = {}

        for jenis, sheet_name in SHEET_MAPPING.items():
            if sheet_name not in wb.sheetnames:
                print(f"[Check] [{jenis}] Sheet tidak ditemukan.")
                results[jenis] = False
                continue

            ws        = wb[sheet_name]
            tahun_col = None
            for col in range(1, ws.max_column + 1):
                if ws.cell(1, col).value == "tahun":
                    tahun_col = col
                    break

            if tahun_col is None:
                print(f"[Check] [{jenis}] Kolom 'tahun' tidak ditemukan.")
                results[jenis] = False
                continue

            found = any(
                str(ws.cell(row, tahun_col).value) == str(tahun)
                for row in range(2, ws.max_row + 1)
                if ws.cell(row, tahun_col).value is not None
            )
            results[jenis] = found
            print(f"[Check] [{jenis}] Data tahun {tahun}: {'Sudah ada' if found else 'Belum ada'}.")

        wb.close()
        return results

    except Exception as exc:
        print(f"[Check] Error saat cek tahun: {exc}")
        traceback.print_exc()
        return default


# Data Fetching

def fetch_all_data(tahun: str = "2025") -> dict[str, pd.DataFrame]:
    """
    Fetch WTE data for all three jenis (sumber, komposisi, timbulan) from SIPSN API.

    Returns a dict {jenis: DataFrame}.
    """
    all_data = {}

    # REVISI 3: Gunakan session agar cookie XSRF-TOKEN dari homepage
    # otomatis tersimpan dan disertakan di setiap POST berikutnya.
    # Tanpa ini, domain baru mengembalikan 403 Forbidden.
    session = requests.Session()
    try:
        session.get(
            "https://sampahnasional.kemenlh.go.id/indikatif",
            headers={k: v for k, v in HEADERS.items()
                     if k not in ["Content-Type", "X-Requested-With"]},
            timeout=15,
        )
    except Exception as exc:
        print(f"[Fetch] Peringatan: gagal ambil homepage untuk cookie — {exc}")

    for jenis in SHEET_MAPPING:
        print(f"[Fetch] Mengambil data {jenis} tahun {tahun}...")
        try:
            response = session.post(          # pakai session, bukan requests.post
                API_URL,
                data={"length": "-1", "jenis": jenis, "tahun": tahun},
                headers=HEADERS,
                timeout=30,
            )
            if response.status_code == 200:
                df = pd.DataFrame(response.json()["data"])
                df = df.rename(columns=COLUMN_RENAME)
                all_data[jenis] = df
                print(f"[Fetch] {jenis}: {len(df)} records.")
            else:
                print(f"[Fetch] Gagal ambil data {jenis}: status {response.status_code}.")
        except Exception as exc:
            print(f"[Fetch] Error {jenis}: {exc}")

    return all_data


# Data Merge — tidak ada perubahan

def _merge_data_by_year(df_existing: pd.DataFrame, df_new: pd.DataFrame) -> pd.DataFrame:
    """
    Concatenate existing and new DataFrames, deduplicate by tahun+location, and sort by tahun.
    """
    df_combined = pd.concat([df_existing, df_new], ignore_index=True)

    id_cols = [c for c in DEDUP_ID_COLS if c in df_combined.columns]
    if id_cols:
        df_combined = df_combined.drop_duplicates(subset=id_cols, keep="last")

    if "tahun" in df_combined.columns:
        df_combined["tahun"] = pd.to_numeric(df_combined["tahun"], errors="coerce")
        df_combined = df_combined.sort_values("tahun", ascending=True)

    print(f"[Merge] Data setelah merge: {len(df_combined)} rows.")
    return df_combined


# Save to OneDrive — tidak ada perubahan

def save_to_onedrive(access_token, data_dict: dict[str, pd.DataFrame], tahun: str) -> None:
    """
    Merge new WTE data with existing OneDrive sheets, deduplicate, and upload.

    Preserves all other sheets in the workbook.
    """
    if not data_dict or all(df.empty for df in data_dict.values()):
        print("[Save] Tidak ada data untuk disimpan.")
        return

    print(f"\n{'='*60}")
    print(f"[Save] Menyimpan data tahun {tahun} ke OneDrive...")
    print(f"{'='*60}")

    excel_buffer  = download_excel_from_onedrive(access_token, ONEDRIVE_FILE_PATH)
    output_buffer = BytesIO()

    try:
        if excel_buffer is None:
            print("[Save] File belum ada — membuat baru...")
            with pd.ExcelWriter(output_buffer, engine="openpyxl", mode="w") as writer:
                for jenis, df in data_dict.items():
                    df.to_excel(writer, sheet_name=SHEET_MAPPING[jenis], index=False)
        else:
            print("[Save] File ditemukan di OneDrive — merging data...")
            excel_buffer.seek(0)
            wb = load_workbook(excel_buffer)

            # Fix hidden sheets
            visible_sheets = [s for s in wb.worksheets if s.sheet_state == "visible"]
            if len(visible_sheets) == 0:
                wb.worksheets[0].sheet_state = "visible"
                wb.active = 0

            for jenis, df_new in data_dict.items():
                sheet_name = SHEET_MAPPING[jenis]
                print(f"\n[Save] Sheet: {sheet_name}")

                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]

                    # Find first non-empty column
                    first_col = 1
                    while first_col <= ws.max_column and ws.cell(1, first_col).value is None:
                        first_col += 1

                    headers = [ws.cell(1, col).value for col in range(first_col, ws.max_column + 1)]
                    rows    = [
                        [ws.cell(row, col).value for col in range(first_col, ws.max_column + 1)]
                        for row in range(2, ws.max_row + 1)
                    ]
                    df_existing = pd.DataFrame(rows, columns=headers)
                    print(f"[Save]   Data existing         : {len(df_existing)} rows.")
                    print(f"[Save]   Data baru (tahun {tahun}): {len(df_new)} rows.")

                    df_merged = _merge_data_by_year(df_existing, df_new)
                    del wb[sheet_name]
                    ws = wb.create_sheet(sheet_name)

                    for col_idx, col_name in enumerate(df_merged.columns, 1):
                        ws.cell(row=1, column=col_idx, value=col_name)
                    for row_idx, row_data in enumerate(df_merged.values, 2):
                        for col_idx, value in enumerate(row_data, 1):
                            ws.cell(row=row_idx, column=col_idx, value=value)
                    print(f"[Save]   Merged: {len(df_merged)} rows.")

                else:
                    print(f"[Save]   Sheet belum ada — membuat baru...")
                    ws = wb.create_sheet(sheet_name)
                    for col_idx, col_name in enumerate(df_new.columns, 1):
                        ws.cell(row=1, column=col_idx, value=col_name)
                    for row_idx, row_data in enumerate(df_new.values, 2):
                        for col_idx, value in enumerate(row_data, 1):
                            ws.cell(row=row_idx, column=col_idx, value=value)
                    print(f"[Save]   Created: {len(df_new)} rows.")

            wb.save(output_buffer)
            wb.close()

        output_buffer.seek(0)

        # Verifikasi
        verify_wb = load_workbook(output_buffer)
        print(f"\n[Save] Verifikasi sheet: {verify_wb.sheetnames}")
        verify_wb.close()
        output_buffer.seek(0)

        # Upload
        print(f"\n[Save] Uploading ke OneDrive: {ONEDRIVE_FILE_PATH}")
        upload_excel_to_onedrive(access_token, ONEDRIVE_FILE_PATH, output_buffer)

        print(f"\n{'='*60}")
        print("[Save] DATA BERHASIL DISIMPAN KE ONEDRIVE")
        print(f"{'='*60}")
        print(f"[Save] File: {ONEDRIVE_FILE_PATH}")

    except Exception as exc:
        print(f"[Save] Error saat menyimpan: {exc}")
        traceback.print_exc()


# Public Entry Point — tidak ada perubahan

def main_sipsn_scraper() -> None:
    print(f"\n{'='*60}")
    print("SCRAPER SIPSN WTE")
    print("STORAGE MODE: OneDrive")
    print(f"{'='*60}")
    print(f"\n[Main] File: {ONEDRIVE_FILE_PATH}")

    print("\n[Main] Authenticating to Microsoft Graph API...")
    try:
        access_token = get_access_token()
        print("[Main] Authentication successful.")
    except Exception as exc:
        print(f"[Main] Authentication failed: {exc}")
        return

    tahun_sekarang = datetime.now().year
    tahun_awal     = 2015

    print(f"\n[Main] Scraping tahun {tahun_awal} s.d. {tahun_sekarang}...")

    for tahun in range(tahun_awal, tahun_sekarang + 1):
        tahun_str = str(tahun)
        print(f"\n{'='*60}")
        print(f"[Main] Memproses tahun {tahun_str}...")
        print(f"{'='*60}")

        data_status = check_year_exists_in_onedrive(access_token, tahun_str)

        if all(data_status.values()):
            print(f"[Main] Semua data tahun {tahun_str} sudah ada — skip.")
            continue

        missing = [jenis for jenis, exists in data_status.items() if not exists]
        print(f"[Main] Data yang belum ada: {', '.join(missing)}")

        data_dict = fetch_all_data(tahun=tahun_str)

        if not data_dict:
            print(f"[Main] Tidak ada data untuk tahun {tahun_str} — kemungkinan belum tersedia, skip.")
            continue

        save_to_onedrive(access_token, data_dict, tahun_str)

    print(f"\n{'='*60}")
    print("[Main] SELESAI!")
    print(f"{'='*60}\n")


# Script Entry Point

if __name__ == "__main__":
    main_sipsn_scraper()