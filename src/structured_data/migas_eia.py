import requests
import pandas as pd
from datetime import datetime
import os
import sys
from io import BytesIO
from openpyxl import load_workbook
from dotenv import load_dotenv
from bs4 import BeautifulSoup
import re 

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from helpers.onedrive_helper import (
    get_access_token,
    download_excel_from_onedrive,
    upload_excel_to_onedrive
)

load_dotenv()

_API_KEY = os.getenv("EIA_API_KEY")
_BASE_API_URL = "https://api.eia.gov/v2/steo/data/"
ONEDRIVE_FILE_PATH = os.getenv("ONEDRIVE_DATA_PATH", "/results/(Terstruktur)Data Scraping.xlsx")
_SHEET_NAME = "(Data)eia"
_STEO_URL = "https://www.eia.gov/outlooks/steo/data/browser/"

_MONTH_TO_NUMBER = {
    'januari': 1, 'februari': 2, 'maret': 3, 'april': 4,
    'mei': 5, 'juni': 6, 'juli': 7, 'agustus': 8,
    'september': 9, 'oktober': 10, 'november': 11, 'desember': 12
}
_NUMBER_TO_MONTH = {
    1: 'Januari', 2: 'Februari', 3: 'Maret', 4: 'April',
    5: 'Mei', 6: 'Juni', 7: 'Juli', 8: 'Agustus',
    9: 'September', 10: 'Oktober', 11: 'November', 12: 'Desember'
}

_SERIES_IDS = {
    'PAPR_WORLD': 'World Total Production',
    'PAPR_OPEC': 'OPEC Production',
    'PAPR_NONOPEC': 'Non-OPEC Production',
    'COPR_WORLD': 'Crude Oil',
    'PATC_WORLD': 'World Total Consumption',
    'PATC_OECD': 'OECD Consumption'
}

def get_eia_release_dates():
    try:
        response = requests.get(_STEO_URL, timeout=10)
        response.raise_for_status()
        soup = BeautifulSoup(response.text, 'html.parser')
        pub_title_div = soup.find('div', class_='pub_title')
        if not pub_title_div:
            return {'success': False, 'error': 'Could not find pub_title div'}
        p_tag = pub_title_div.find('p')
        if not p_tag:
            return {'success': False, 'error': 'Could not find paragraph tag'}
        text = p_tag.get_text()
        release_match = re.search(r'Release Date:\s*(\w+)\s+(\d+),\s+(\d+)', text)
        if not release_match:
            return {'success': False, 'error': 'Could not parse Release Date'}
        next_match = re.search(r'Next Release Date:\s*(\w+)\s+(\d+),\s+(\d+)', text)
        if not next_match:
            return {'success': False, 'error': 'Could not parse Next Release Date'}
        month_map = {
            'January': 1, 'February': 2, 'March': 3, 'April': 4,
            'May': 5, 'June': 6, 'July': 7, 'August': 8,
            'September': 9, 'October': 10, 'November': 11, 'December': 12
        }
        release_month_name = release_match.group(1)
        release_day = int(release_match.group(2))
        release_year = int(release_match.group(3))
        release_month = month_map.get(release_month_name)
        if not release_month:
            return {'success': False, 'error': f'Unknown month: {release_month_name}'}
        release_date = datetime(release_year, release_month, release_day)
        release_date_str = f"{release_month_name} {release_day}, {release_year}"
        next_month_name = next_match.group(1)
        next_day = int(next_match.group(2))
        next_year = int(next_match.group(3))
        next_month = month_map.get(next_month_name)
        if not next_month:
            return {'success': False, 'error': f'Unknown month: {next_month_name}'}
        next_release_date = datetime(next_year, next_month, next_day)
        next_release_date_str = f"{next_month_name} {next_day}, {next_year}"
        return {
            'success': True,
            'release_date': release_date,
            'release_date_str': release_date_str,
            'next_release_date': next_release_date,
            'next_release_date_str': next_release_date_str
        }    
    except requests.exceptions.RequestException as e:
        return {'success': False, 'error': f'Network error: {e}'}
    except Exception as e:
        return {'success': False, 'error': f'Unexpected error: {e}'}


def read_last_entry_from_excel(access_token, sheet_name):
    try:
        excel_buffer = download_excel_from_onedrive(access_token, ONEDRIVE_FILE_PATH)
        if excel_buffer is None:
            print("File scraping tidak ditemukan di OneDrive. Semua data akan diunduh.")
            return None, None
        df = pd.read_excel(excel_buffer, sheet_name=sheet_name, engine='openpyxl')
        if df.empty or "Bulan" not in df.columns or "Tahun" not in df.columns:
            print("Sheet kosong atau format salah. Semua data akan diunduh.")
            return None, None
        df["Bulan"] = df["Bulan"].astype(str).str.lower()
        df["Bulan_Angka"] = df["Bulan"].map(_MONTH_TO_NUMBER)
        df = df.dropna(subset=["Bulan_Angka"])
        if df.empty:
            print("Tidak ada data valid di Excel. Semua data akan diunduh.")
            return None, None
        df_sorted = df.sort_values(["Tahun", "Bulan_Angka"])
        last_row = df_sorted.iloc[-1]
        print(f"Data terakhir di Excel: {last_row['Bulan'].capitalize()} {int(last_row['Tahun'])}")
        return int(last_row["Tahun"]), int(last_row["Bulan_Angka"])
    except ValueError:
        print(f"Sheet '{sheet_name}' tidak ditemukan. Semua data akan diunduh.")
        return None, None
    except Exception as e:
        print(f"Error membaca Excel: {e}")
        return None, None

def get_migas_eia_needed_data(access_token):
    last_read_year, last_read_month = read_last_entry_from_excel(access_token, _SHEET_NAME)
    today = datetime.today()
    month = today.month - 1
    year = today.year
    if month == 0:
        month = 12
        year -= 1
    needed_data = []
    if last_read_year is None or last_read_month is None:
        print(" No existing data found. Will fetch all available data.")
        return None
    next_month = last_read_month + 1
    next_year = last_read_year
    if next_month > 12:
        next_month = 1
        next_year += 1
    if next_year > year or (next_year == year and next_month > month):
        print(f"Data already up to date. Last data: {_NUMBER_TO_MONTH[last_read_month]} {last_read_year}")
        return []
    needed_data.append((next_year, next_month))
    print(f"Will fetch 1 new month: {_NUMBER_TO_MONTH[next_month]} {next_year}")
    return needed_data

def get_start_end_from_needed_data(needed_data):
    if needed_data is None:
        today = datetime.today()
        end_year = today.year
        end_month = today.month - 1
        if end_month == 0:
            end_month = 12
            end_year -= 1
        start_str = "2015-01"
        end_str = f"{end_year}-{end_month:02d}"
        print(f"Fetching all data from {start_str} to {end_str}")
        return start_str, end_str
    if not needed_data:
        return None, None
    year, month = needed_data[0]
    date_str = f"{year}-{month:02d}"
    print(f"Fetching data for: {date_str}")
    return date_str, date_str

def fetch_eia_data(series_id, start, end):
    params = {
        "api_key": _API_KEY,
        "frequency": "monthly",
        "data[0]": "value",
        "facets[seriesId][]": series_id,
        "start": start,
        "end": end
    }
    print(f"  Fetching {series_id}...")
    try:
        response = requests.get(_BASE_API_URL, params=params, timeout=30)
        response.raise_for_status()
        data = response.json()
        if "response" in data and "data" in data["response"]:
            records = data["response"]["data"]
            print(f"Got {len(records)} records")
            return records
        else:
            print(f"No data found for {series_id}")
            return []
    except requests.exceptions.RequestException as e:
        print(f"Error fetching {series_id}: {e}")
        return []
    
def should_run_scraping(access_token):
    print("Checking if scraping is needed...")
    release_info = get_eia_release_dates()
    if not release_info['success']:
        print(f"Warning: Could not get release dates from website: {release_info.get('error')}")
        print("Will proceed with scraping anyway...")
        return True, None
    print(f"Current Release Date from website: {release_info['release_date_str']}")
    print(f"Current Next Release Date from website: {release_info['next_release_date_str']}")
    try:
        excel_buffer = download_excel_from_onedrive(access_token, ONEDRIVE_FILE_PATH)
        if excel_buffer is None:
            print("No existing file. Will run scraping.")
            return True, release_info
        df = pd.read_excel(excel_buffer, sheet_name=_SHEET_NAME, engine='openpyxl')
        if df.empty or 'Next Release Date' not in df.columns:
            print("No Next Release Date found in Excel. Will run scraping.")
            return True, release_info
        last_next_release_str = df['Next Release Date'].iloc[-1]
        if pd.isna(last_next_release_str):
            print("Next Release Date is empty. Will run scraping.")
            return True, release_info
        last_next_release = pd.to_datetime(last_next_release_str).replace(tzinfo=None)
        print(f"Last Next Release Date from Excel: {last_next_release.strftime('%B %d, %Y')}")
        today = datetime.today().replace(hour=0, minute=0, second=0, microsecond=0)
        if today > last_next_release:
            print(f"Today ({today.strftime('%Y-%m-%d')}) >= Next Release Date ({last_next_release.strftime('%Y-%m-%d')})")
            print("New data should be available. Will run scraping.")
            return True, release_info
        else:
            print(f"Today ({today.strftime('%Y-%m-%d')}) < Next Release Date ({last_next_release.strftime('%Y-%m-%d')})")
            print("No new data yet. Skipping scraping.")
            return False, release_info      
    except ValueError:
        print(f"Sheet '{_SHEET_NAME}' not found. Will run scraping.")
        return True, release_info
    except Exception as e:
        print(f"Error checking Next Release Date: {e}")
        print("Will proceed with scraping anyway...")
        return True, release_info

def transform_data_to_excel_format(all_data, release_info=None):
    period_data = {}
    for series_id, records in all_data.items():
        for record in records:
            period = record.get('period')
            value = record.get('value')
            if period and value and value != 'w':
                if period not in period_data:
                    period_data[period] = {}
                try:
                    period_data[period][series_id] = round(float(value), 2)
                except (ValueError, TypeError):
                    period_data[period][series_id] = None
    next_release_str = None
    if release_info and release_info['success']:
        next_release_str = release_info['next_release_date_str']
    rows = []
    for period in sorted(period_data.keys()):
        year, month = period.split('-')
        year = int(year)
        month = int(month)
        world_total_prod = period_data[period].get('PAPR_WORLD')
        opec = period_data[period].get('PAPR_OPEC')
        non_opec = period_data[period].get('PAPR_NONOPEC')
        crude_oil = period_data[period].get('COPR_WORLD')
        world_total_cons = period_data[period].get('PATC_WORLD')
        oecd = period_data[period].get('PATC_OECD')
        
        world_total_prod = round(world_total_prod, 2) if world_total_prod is not None else None
        opec = round(opec, 2) if opec is not None else None
        non_opec = round(non_opec, 2) if non_opec is not None else None
        crude_oil = round(crude_oil, 2) if crude_oil is not None else None
        world_total_cons = round(world_total_cons, 2) if world_total_cons is not None else None
        oecd = round(oecd, 2) if oecd is not None else None
        
        other_liquids = None
        if world_total_prod is not None and crude_oil is not None:
            other_liquids = round(world_total_prod - crude_oil, 2)
        
        non_oecd = None
        if world_total_cons is not None and oecd is not None:
            non_oecd = round(world_total_cons - oecd, 2)
        
        row = {
            'Bulan': _NUMBER_TO_MONTH.get(month, f'Month-{month}'),
            'Tahun': year,
            'Next Release Date': next_release_str,
            'World Total Production': world_total_prod,
            'OPEC': opec,
            'Non-OPEC': non_opec,
            'Crude Oil': crude_oil,
            'Other Liquids': other_liquids,
            'World Total Consumption': world_total_cons,
            'OECD': oecd,
            'Non-OECD': non_oecd
        }
        rows.append(row)
    return pd.DataFrame(rows)

def save_to_excel(access_token, df, sheet_name):
    print("Step 5: Saving to Excel...")
    if df.empty:
        print("DataFrame kosong, tidak ada yang disimpan")
        return
    excel_buffer = download_excel_from_onedrive(access_token, ONEDRIVE_FILE_PATH)
    if excel_buffer is None:
        print("File tidak ada di OneDrive, akan membuat baru")
        df_combined = df
    else:
        try:
            existing_df = pd.read_excel(excel_buffer, sheet_name=sheet_name, engine='openpyxl')
            df_combined = pd.concat([existing_df, df], ignore_index=True)
            df_combined.drop_duplicates(subset=['Bulan', 'Tahun'], keep='last', inplace=True)
            month_order = {
                'Januari': 1, 'Februari': 2, 'Maret': 3, 'April': 4,
                'Mei': 5, 'Juni': 6, 'Juli': 7, 'Agustus': 8,
                'September': 9, 'Oktober': 10, 'November': 11, 'Desember': 12
            }
            df_combined['Bulan_Order'] = df_combined['Bulan'].map(month_order)
            df_combined = df_combined.sort_values(['Tahun', 'Bulan_Order'])
            df_combined = df_combined.drop('Bulan_Order', axis=1)
            print(f"Merged with existing data. Total rows: {len(df_combined)}")
        except ValueError:
            print(f"Sheet '{sheet_name}' tidak ditemukan, membuat sheet baru")
            df_combined = df
        except Exception as e:
            print(f"Error reading existing file: {e}")
            print("Creating new file instead...")
            df_combined = df
    output_buffer = BytesIO()
    try:
        if excel_buffer is None:
            print("Membuat file Excel baru...")
            with pd.ExcelWriter(output_buffer, engine='openpyxl', mode='w') as writer:
                df_combined.to_excel(writer, sheet_name=sheet_name, index=False)
        else:
            print("File existing ditemukan, preserve semua sheet...")
            excel_buffer.seek(0)
            wb = load_workbook(excel_buffer)
            print(f"Sheet yang ada saat ini: {wb.sheetnames}")
            visible_sheets = [s for s in wb.worksheets if s.sheet_state == 'visible']
            if len(visible_sheets) == 0:
                print("Fixing hidden sheets...")
                wb.worksheets[0].sheet_state = 'visible'
                wb.active = 0
            for sheet in wb.worksheets:
                if sheet.sheet_state != 'visible':
                    sheet.sheet_state = 'visible'
            if sheet_name in wb.sheetnames:
                print(f"Menghapus sheet '{sheet_name}' yang lama...")
                del wb[sheet_name]
            ws = wb.create_sheet(sheet_name)
            for col_idx, col_name in enumerate(df_combined.columns, 1):
                ws.cell(row=1, column=col_idx, value=col_name)
            for row_idx, row_data in enumerate(df_combined.values, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            print(f"Sheet yang akan disimpan: {wb.sheetnames}")
            wb.save(output_buffer)
            wb.close()
        output_buffer.seek(0)
        verify_wb = load_workbook(output_buffer)
        print(f"Verifikasi - Sheet di buffer: {verify_wb.sheetnames}")
        verify_wb.close()
        output_buffer.seek(0)
        print(f"\nUploading ke OneDrive: {ONEDRIVE_FILE_PATH}")
        upload_excel_to_onedrive(access_token, ONEDRIVE_FILE_PATH, output_buffer)
        print("Upload selesai!")
        print(f"Saved to OneDrive: {ONEDRIVE_FILE_PATH}")
        print(f"Sheet: {sheet_name}")
        print(f"Total records: {len(df_combined)}")
    except Exception as e:
        print(f"Error saving to Excel: {e}")
        import traceback
        traceback.print_exc()

def main_eia():
    print("=" * 80)
    print("EIA STEO DATA SCRAPER")
    print("STORAGE MODE: OneDrive")
    print("=" * 80)
    print()
    print(f"File: {ONEDRIVE_FILE_PATH}")
    print(f"Sheet: {_SHEET_NAME}")
    print("\nAuthenticating to Microsoft Graph API...")
    try:
        access_token = get_access_token()
        print("Authentication successful")
    except Exception as e:
        print(f"Authentication failed: {e}")
        return
    print()
    should_run, release_info = should_run_scraping(access_token)
    if not should_run:
        print("\n" + "=" * 80)
        print("SKIPPED: No update needed yet.")
        print("=" * 80)
        return
    print()
    print("Step 1: Checking what data needs to be downloaded...")
    needed_data = get_migas_eia_needed_data(access_token)
    if needed_data is not None and len(needed_data) == 0:
        print("All data is up to date!")
        return
    if needed_data is None:
        print("Will download ALL available data from EIA STEO")
    else:
        print(f"Need to download data for {len(needed_data)} months")
        print(f"Periods: {needed_data[0]} to {needed_data[-1]}")
    print()
    start, end = get_start_end_from_needed_data(needed_data)
    print(f"Step 2: Date range: {start} to {end}")
    print()
    print("Step 3: Fetching data from EIA API...")
    all_data = {}
    for series_id in _SERIES_IDS.keys():
        records = fetch_eia_data(series_id, start, end)
        if records:
            all_data[series_id] = records
    if not all_data:
        print("\nNo data fetched. Please check your API key and internet connection.")
        return
    print()
    print("Step 4: Transforming data to Excel format...")
    df = transform_data_to_excel_format(all_data, release_info)
    if df.empty:
        print("No valid data to save.")
        return
    print(f"Transformed {len(df)} rows")
    print()
    print("Preview of data:")
    print(df.head(10).to_string(index=False))
    print()
    print("Step 5: Saving to OneDrive...")
    save_to_excel(access_token, df, _SHEET_NAME)
    print()
    print("=" * 80)
    print("DONE!")
    print("=" * 80)

if __name__ == "__main__":
    print("=" * 80)
    print("EIA STEO DATA SCRAPER")
    print("STORAGE MODE: OneDrive")
    print("=" * 80)
    print()
    print(f"File: {ONEDRIVE_FILE_PATH}")
    print(f"Sheet: {_SHEET_NAME}")
    print("\nAuthenticating to Microsoft Graph API...")
    try:
        access_token = get_access_token()
        print("Authentication successful")
    except Exception as e:
        print(f"Authentication failed: {e}")
    should_run, release_info = should_run_scraping(access_token)
    needed_data = get_migas_eia_needed_data(access_token)
    print(needed_data)