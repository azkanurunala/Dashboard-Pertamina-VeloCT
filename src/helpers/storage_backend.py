"""
Storage backend abstraction layer.

Switch between OneDrive/Excel and Neon PostgreSQL by setting:
    STORAGE_BACKEND=onedrive   (default, uses MS Graph API)
    STORAGE_BACKEND=neon       (uses psycopg2 + NEON_DB_URL)

All orchestrators and structured_data scripts import `storage` from here.
onedrive_helper.py is kept unchanged and used internally by _OneDriveBackend.
"""

import os
from io import BytesIO

import pandas as pd
from dotenv import load_dotenv

load_dotenv()

STORAGE_BACKEND: str = os.getenv("STORAGE_BACKEND", "onedrive")

_NEWS_FILE_PATH      = os.getenv("ONEDRIVE_FILE_PATH",      "/results/(News)Scrapping.xlsx")
_SENTIMENT_FILE_PATH = os.getenv("ONEDRIVE_SENTIMENT_PATH", "/results/(News)Sentiment.xlsx")
_DATA_FILE_PATH      = os.getenv("ONEDRIVE_DATA_PATH",      "/results/(Terstruktur)Data Scrapping.xlsx")

# Excel sheet → PostgreSQL table name
SHEET_TO_TABLE: dict[str, str] = {
    "(Data)Biodesel":              "data_biodiesel",
    "(Data)Bioetanol":             "data_bioetanol",
    "(Data)Harga Minyak":          "data_harga_minyak",
    "(Data)eia":                   "data_eia",
    "(Data)CPO":                   "data_cpo",
    "(Data)SAF":                   "data_saf",
    "(Data)Kapasitas_EBT":         "data_kapasitas_ebt",
    "(Data)WTE_Sumber":            "data_wte_sumber",
    "(Data)WTE_Komposisi":         "data_wte_komposisi",
    "(Data)WTE_Timbulan":          "data_wte_timbulan",
    "(Data)IAEA_Nuclear_Capacity": "data_iaea_nuclear_capacity",
    "(Data)IAEA_Electrical":       "data_iaea_electrical",
    "(Data)IAEA_Country_Stats":    "data_iaea_country_stats",
    "(Data)Crackspread_BBM":       "data_crackspread_bbm",
    "(Data)Crackspread_NON_BBM":   "data_crackspread_non_bbm",
    "(Data)Crackspread_BBM_YEAR":  "data_crackspread_bbm_year",
}

# Conflict columns used for ON CONFLICT upsert
SHEET_CONFLICT_COLS: dict[str, list[str]] = {
    "(Data)Biodesel":              ["Bulan HIP"],
    "(Data)Bioetanol":             ["Bulan HIP"],
    "(Data)Harga Minyak":          ["Tahun", "Bulan"],
    "(Data)eia":                   ["Tahun", "Bulan"],
    "(Data)CPO":                   ["Dates"],
    "(Data)SAF":                   ["assessDate"],
    "(Data)Kapasitas_EBT":         ["tahun", "bulan"],
    "(Data)WTE_Sumber":            ["tahun", "Nama Provinsi", "Nama Kota/Kabupaten"],
    "(Data)WTE_Komposisi":         ["tahun", "Nama Provinsi", "Nama Kota/Kabupaten"],
    "(Data)WTE_Timbulan":          ["tahun", "Nama Provinsi", "Nama Kota/Kabupaten"],
    "(Data)IAEA_Nuclear_Capacity": ["year", "country"],
    "(Data)IAEA_Electrical":       ["year", "country"],
    "(Data)IAEA_Country_Stats":    ["CountryCode"],
    "(Data)Crackspread_BBM":       ["year", "month"],
    "(Data)Crackspread_NON_BBM":   ["Year", "Month"],
    "(Data)Crackspread_BBM_YEAR":  ["year"],
}

# IAEA sheets that need wide↔long transform (rows=years, cols=countries in Excel)
_IAEA_VALUE_COL: dict[str, str] = {
    "(Data)IAEA_Nuclear_Capacity": "value_mw",
    "(Data)IAEA_Electrical":       "value_twh",
}


def _melt_iaea(df: pd.DataFrame, sheet_name: str) -> pd.DataFrame:
    """Wide (Year col + country cols) → long (year, country, value)."""
    value_col   = _IAEA_VALUE_COL[sheet_name]
    country_cols = [c for c in df.columns if c != "Year"]
    melted = df.melt(id_vars=["Year"], value_vars=country_cols,
                     var_name="country", value_name=value_col)
    melted = melted.rename(columns={"Year": "year"})
    melted["year"] = pd.to_numeric(melted["year"], errors="coerce")
    return melted.dropna(subset=[value_col]).reset_index(drop=True)


def _pivot_iaea(df: pd.DataFrame, sheet_name: str) -> pd.DataFrame:
    """Long (year, country, value) → wide (Year col + country cols)."""
    value_col = _IAEA_VALUE_COL[sheet_name]
    df = df.drop(columns=["id"], errors="ignore")
    pivoted = df.pivot_table(
        index="year", columns="country", values=value_col, aggfunc="first"
    ).reset_index()
    pivoted.columns.name = None
    return pivoted.rename(columns={"year": "Year"})


class _OneDriveBackend:
    def __init__(self) -> None:
        self._cached_token: str | None = None

    def _token(self) -> str:
        if self._cached_token is None:
            from helpers.onedrive_helper import get_access_token
            self._cached_token = get_access_token()
        return self._cached_token

    def _refresh(self) -> str:
        from helpers.onedrive_helper import get_access_token
        self._cached_token = get_access_token()
        return self._cached_token

    # ── News ────────────────────────────────────────────────────────────────

    def read_news_sheet(self, sheet_name: str) -> pd.DataFrame:
        from helpers.onedrive_helper import read_excel_sheet_from_onedrive
        return read_excel_sheet_from_onedrive(self._token(), _NEWS_FILE_PATH, sheet_name)

    def read_all_news_sheets(self, sheet_names: list[str]) -> dict[str, pd.DataFrame]:
        from helpers.onedrive_helper import download_excel_from_onedrive
        buf = download_excel_from_onedrive(self._token(), _NEWS_FILE_PATH)
        if buf is None:
            return {s: pd.DataFrame() for s in sheet_names}
        xf = pd.ExcelFile(buf)
        result = {}
        for s in sheet_names:
            try:
                result[s] = pd.read_excel(xf, sheet_name=s) if s in xf.sheet_names else pd.DataFrame()
            except Exception:
                result[s] = pd.DataFrame()
        xf.close()
        return result

    def write_news_file(self, all_sheets: dict[str, pd.DataFrame]) -> None:
        from helpers.onedrive_helper import write_multiple_sheets_to_onedrive
        write_multiple_sheets_to_onedrive(self._refresh(), _NEWS_FILE_PATH, all_sheets)

    # ── Sentiment ───────────────────────────────────────────────────────────

    def read_sentiment_sheet(self, sheet_name: str) -> pd.DataFrame:
        from helpers.onedrive_helper import read_excel_sheet_from_onedrive
        return read_excel_sheet_from_onedrive(self._token(), _SENTIMENT_FILE_PATH, sheet_name)

    def read_all_sentiment_sheets(self, sheet_names: list[str]) -> dict[str, pd.DataFrame]:
        from helpers.onedrive_helper import download_excel_from_onedrive
        buf = download_excel_from_onedrive(self._token(), _SENTIMENT_FILE_PATH)
        if buf is None:
            return {s: pd.DataFrame() for s in sheet_names}
        xf = pd.ExcelFile(buf)
        result = {}
        for s in sheet_names:
            try:
                result[s] = pd.read_excel(xf, sheet_name=s) if s in xf.sheet_names else pd.DataFrame()
            except Exception:
                result[s] = pd.DataFrame()
        xf.close()
        return result

    def write_sentiment_file(self, all_sheets: dict[str, pd.DataFrame]) -> None:
        from helpers.onedrive_helper import write_multiple_sheets_to_onedrive
        write_multiple_sheets_to_onedrive(self._refresh(), _SENTIMENT_FILE_PATH, all_sheets)

    # ── Structured data ─────────────────────────────────────────────────────

    def read_structured_sheet(self, sheet_name: str) -> pd.DataFrame:
        from helpers.onedrive_helper import read_excel_sheet_from_onedrive
        return read_excel_sheet_from_onedrive(self._token(), _DATA_FILE_PATH, sheet_name)

    def write_structured_sheet(self, sheet_name: str, df: pd.DataFrame) -> None:
        """Update one sheet in the structured workbook, preserving all others."""
        from openpyxl import load_workbook
        from helpers.onedrive_helper import download_excel_from_onedrive, upload_excel_to_onedrive

        buf    = download_excel_from_onedrive(self._token(), _DATA_FILE_PATH)
        output = BytesIO()

        if buf is None:
            with pd.ExcelWriter(output, engine="openpyxl", mode="w") as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        else:
            wb = load_workbook(buf)
            if not any(s.sheet_state == "visible" for s in wb.worksheets):
                wb.worksheets[0].sheet_state = "visible"
                wb.active = 0
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]
            ws = wb.create_sheet(sheet_name)
            for ci, col in enumerate(df.columns, 1):
                ws.cell(row=1, column=ci, value=col)
            for ri, row in enumerate(df.itertuples(index=False, name=None), 2):
                for ci, val in enumerate(row, 1):
                    ws.cell(row=ri, column=ci, value=val)
            wb.save(output)

        output.seek(0)
        upload_excel_to_onedrive(self._refresh(), _DATA_FILE_PATH, output)


class _NeonBackend:
    # ── News ────────────────────────────────────────────────────────────────

    def read_news_sheet(self, sheet_name: str) -> pd.DataFrame:
        from helpers import neon_helper
        df = neon_helper.read_table("news_articles", topic=sheet_name)
        return df.drop(columns=["id", "topic"], errors="ignore")

    def read_all_news_sheets(self, sheet_names: list[str]) -> dict[str, pd.DataFrame]:
        return {s: self.read_news_sheet(s) for s in sheet_names}

    def write_news_file(self, all_sheets: dict[str, pd.DataFrame]) -> None:
        from helpers import neon_helper
        for sheet_name, df in all_sheets.items():
            if df.empty:
                continue
            df_out = df.copy()
            df_out["topic"] = sheet_name
            neon_helper.upsert_df("news_articles", df_out, ["url", "topic"])

    # ── Sentiment ───────────────────────────────────────────────────────────

    def read_sentiment_sheet(self, sheet_name: str) -> pd.DataFrame:
        from helpers import neon_helper
        df = neon_helper.read_table("news_sentiment", topic=sheet_name)
        return df.drop(columns=["id", "topic"], errors="ignore")

    def read_all_sentiment_sheets(self, sheet_names: list[str]) -> dict[str, pd.DataFrame]:
        return {s: self.read_sentiment_sheet(s) for s in sheet_names}

    def write_sentiment_file(self, all_sheets: dict[str, pd.DataFrame]) -> None:
        from helpers import neon_helper
        for sheet_name, df in all_sheets.items():
            if df.empty:
                continue
            df_out = df.copy()
            df_out["topic"] = sheet_name
            neon_helper.upsert_df("news_sentiment", df_out, ["topic", "Tanggal awal"])

    # ── Structured data ─────────────────────────────────────────────────────

    def read_structured_sheet(self, sheet_name: str) -> pd.DataFrame:
        from helpers import neon_helper
        table = SHEET_TO_TABLE[sheet_name]
        df    = neon_helper.read_table(table)
        df    = df.drop(columns=["id"], errors="ignore")
        if sheet_name in _IAEA_VALUE_COL:
            df = _pivot_iaea(df, sheet_name)
        return df

    def write_structured_sheet(self, sheet_name: str, df: pd.DataFrame) -> None:
        from helpers import neon_helper
        table    = SHEET_TO_TABLE[sheet_name]
        conflict = SHEET_CONFLICT_COLS[sheet_name]
        if sheet_name in _IAEA_VALUE_COL:
            df = _melt_iaea(df, sheet_name)
        # WTE tables: auto-create/alter table based on DataFrame schema
        wte_sheets = {"(Data)WTE_Sumber", "(Data)WTE_Komposisi", "(Data)WTE_Timbulan"}
        if sheet_name in wte_sheets:
            neon_helper.create_table_if_needed(table, df, conflict)
        neon_helper.upsert_df(table, df, conflict)


# Singleton — import `storage` everywhere instead of onedrive_helper
storage: _OneDriveBackend | _NeonBackend = (
    _NeonBackend() if STORAGE_BACKEND == "neon" else _OneDriveBackend()
)
