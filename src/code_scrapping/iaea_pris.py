import os
import sys
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from dotenv import load_dotenv
import time

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from helpers.onedrive_helper import (
    get_access_token,
    download_excel_from_onedrive,
    upload_excel_to_onedrive
)
from helpers.scraping_helper import setup_driver

load_dotenv()

ONEDRIVE_FILE_PATH = os.getenv("ONEDRIVE_DATA_PATH", "/results/(Terstruktur)Data Scrapping.xlsx")
URL_CAPACITY = "https://pris.iaea.org/PRIS/WorldStatistics/WorldTrendNuclearPowerCapacity.aspx"
URL_PRODUCTION = "https://pris.iaea.org/PRIS/WorldStatistics/WorldTrendinElectricalProduction.aspx"
URL_COUNTRY_STATS = "https://pris.iaea.org/PRIS/CountryStatistics/CountryStatisticsLandingPage.aspx"
SHEET_NAME_CAPACITY = "(Data)IAEA_Nuclear_Capacity"
SHEET_NAME_PRODUCTION = "(Data)IAEA_Electrical"
SHEET_NAME_COUNTRY_STATS = "(Data)IAEA_Country_Stats"

from urllib.parse import urlparse, parse_qs

def _to_number(x: str):
    x = (x or "").strip()
    if x == "":
        return None
    try:
        # data di landing page berupa integer (reaktor) & integer MW
        return int(x)
    except ValueError:
        try:
            return float(x)
        except ValueError:
            return x

def _parse_hash_pipe_line(line: str, col_names: list[str]) -> pd.DataFrame:
    # line: "COUNTRY|a|b|c#COUNTRY2|a|b|c..."
    records = []
    for item in line.split("#"):
        parts = [p.strip() for p in item.split("|")]
        if len(parts) != len(col_names):
            continue
        rec = {"Country": parts[0].strip()}
        for i, c in enumerate(col_names[1:], start=1):
            rec[c] = _to_number(parts[i])
        records.append(rec)
    return pd.DataFrame(records)

def fetch_nuclear_capacity_data():
    driver = None
    try:
        driver = setup_driver()
        driver.get(URL_CAPACITY)
        wait = WebDriverWait(driver, 20)
        wait.until(EC.presence_of_element_located((By.TAG_NAME, "table")))
        time.sleep(3)
        html = driver.page_source
        soup = BeautifulSoup(html, 'html.parser')
        table = soup.find('table')
        if not table:
            print("Tabel tidak ditemukan")
            return None
        headers = []
        thead = table.find('thead')
        if thead:
            for th in thead.find_all('th'):
                text = th.get_text(strip=True).replace('\n', ' ')
                headers.append(text)
        data = []
        tbody = table.find('tbody')
        if tbody:
            for tr in tbody.find_all('tr'):
                row = []
                for td in tr.find_all('td'):
                    text = td.get_text(strip=True)
                    try:
                        if '.' in text:
                            row.append(float(text))
                        else:
                            row.append(int(text))
                    except ValueError:
                        row.append(text)
                if row:
                    data.append(row)
        df = pd.DataFrame(data, columns=headers)
        print(f"Berhasil scrape {len(df)} baris data Nuclear Capacity")
        return df
    except Exception as e:
        print(f"Error: {e}")
        return None
    finally:
        if driver:
            driver.quit()

def fetch_electrical_production_data():
    driver = None
    try:
        driver = setup_driver()
        driver.get(URL_PRODUCTION)
        wait = WebDriverWait(driver, 20)
        wait.until(EC.presence_of_element_located((By.TAG_NAME, "table")))
        time.sleep(3)
        html = driver.page_source
        soup = BeautifulSoup(html, 'html.parser')
        table = soup.find('table')
        if not table:
            print("Tabel tidak ditemukan")
            return None
        headers = []
        thead = table.find('thead')
        if thead:
            for th in thead.find_all('th'):
                text = th.get_text(strip=True).replace('\n', ' ')
                if text: 
                    headers.append(text)
        data = []
        tbody = table.find('tbody')
        if tbody:
            for tr in tbody.find_all('tr'):
                row = []
                for td in tr.find_all('td'):
                    text = td.get_text(strip=True)
                    if not text or text == '\xa0' or text == '':
                        continue
                    try:
                        if '.' in text:
                            row.append(float(text))
                        else:
                            row.append(int(text))
                    except ValueError:
                        row.append(text)
                if row and len(row) == len(headers):
                    data.append(row)
        df = pd.DataFrame(data, columns=headers)
        print(f"Berhasil scrape {len(df)} baris data Electrical Production")
        return df
    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
        return None
    finally:
        if driver:
            driver.quit()

def fetch_country_statistics_data():
    driver = None
    try:
        driver = setup_driver()
        driver.get(URL_COUNTRY_STATS)
        wait = WebDriverWait(driver, 20)
        # tunggu konten utama muncul
        wait.until(EC.presence_of_element_located((By.TAG_NAME, "body")))
        time.sleep(2)

        soup = BeautifulSoup(driver.page_source, "html.parser")

        # 1) ambil 2 baris "string statistik" (reaktor & kapasitas)
        lines = [l.strip() for l in soup.get_text("\n").splitlines() if l.strip()]
        stats_lines = [l for l in lines if ("#" in l and "|" in l)]
        if len(stats_lines) < 2:
            print("Tidak menemukan 2 baris statistik country di landing page")
            return None

        reactors_line = stats_lines[0]
        capacity_line = stats_lines[1]

        df_reactors = _parse_hash_pipe_line(
            reactors_line,
            ["Country", "Reactors_InOperation", "Reactors_UnderConstruction", "Reactors_PermanentShutdown"]
        )
        df_capacity = _parse_hash_pipe_line(
            capacity_line,
            ["Country", "NetCapacityMW_InOperation", "NetCapacityMW_UnderConstruction", "NetCapacityMW_PermanentShutdown"]
        )

        # 2) ambil mapping Country -> CountryCode (parameter current=XX dari link detail negara)
        code_map = {}
        for a in soup.select("a[href*='CountryDetails.aspx?current=']"):
            href = a.get("href", "")
            name = (a.get_text(strip=True) or "").strip()
            if not href or not name:
                continue
            qs = parse_qs(urlparse(href).query)
            cc = (qs.get("current", [None])[0] or "").strip()
            if cc:
                code_map[name.upper()] = cc  # samakan format dengan baris stats (uppercase)

        # normalize country ke uppercase untuk merge aman
        for d in (df_reactors, df_capacity):
            d["Country"] = d["Country"].astype(str).str.strip()

        df = df_reactors.merge(df_capacity, on="Country", how="outer")
        df["CountryKey"] = df["Country"].astype(str).str.upper()
        df["CountryCode"] = df["CountryKey"].map(code_map)
        df.drop(columns=["CountryKey"], inplace=True)

        # opsional: taruh CountryCode di depan
        cols = ["Country", "CountryCode"] + [c for c in df.columns if c not in ("Country", "CountryCode")]
        df = df[cols]

        print(f"Berhasil scrape {len(df)} baris Country Statistics")
        return df

    except Exception as e:
        print(f"Error fetch_country_statistics_data: {e}")
        import traceback
        traceback.print_exc()
        return None
    finally:
        if driver:
            driver.quit()
            
def save_to_onedrive(access_token, df_capacity: pd.DataFrame, df_production: pd.DataFrame, df_country: pd.DataFrame):
    print("\n" + "="*80)
    print("MENYIMPAN HASIL KE ONEDRIVE")
    print("="*80)
    if (df_capacity is None or df_capacity.empty) and \
       (df_production is None or df_production.empty) and \
       (df_country is None or df_country.empty):
        print("Tidak ada data untuk disimpan")
        return
    excel_buffer = download_excel_from_onedrive(access_token, ONEDRIVE_FILE_PATH)
    output_buffer = BytesIO()
    try:
        if excel_buffer is None:
            print("File tidak ada di OneDrive, membuat file baru...")
            wb = load_workbook(BytesIO())
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']
        else:
            print("File ditemukan di OneDrive, updating...")
            excel_buffer.seek(0)
            wb = load_workbook(excel_buffer)
            visible_sheets = [s for s in wb.worksheets if s.sheet_state == 'visible']
            if len(visible_sheets) == 0:
                wb.worksheets[0].sheet_state = 'visible'
                wb.active = 0
        if df_capacity is not None and not df_capacity.empty:
            print(f"\nMemproses sheet: {SHEET_NAME_CAPACITY}")
            df_combined_capacity = df_capacity
            if excel_buffer is not None:
                excel_buffer.seek(0)
                try:
                    existing_df = pd.read_excel(excel_buffer, sheet_name=SHEET_NAME_CAPACITY, engine='openpyxl')
                    df_combined_capacity = pd.concat([existing_df, df_capacity], ignore_index=True)
                    df_combined_capacity.drop_duplicates(subset=["Year"], keep="last", inplace=True)
                    df_combined_capacity = df_combined_capacity.sort_values("Year")
                    print(f"  Data di-merge. Total rows: {len(df_combined_capacity)}")
                except:
                    print(f"  Sheet baru akan dibuat")
            if SHEET_NAME_CAPACITY in wb.sheetnames:
                del wb[SHEET_NAME_CAPACITY]
            ws = wb.create_sheet(SHEET_NAME_CAPACITY)
            for col_idx, col_name in enumerate(df_combined_capacity.columns, 1):
                ws.cell(row=1, column=col_idx, value=col_name)
            for row_idx, row_data in enumerate(df_combined_capacity.values, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            print(f"Sheet '{SHEET_NAME_CAPACITY}': {len(df_combined_capacity)} rows")
        if df_production is not None and not df_production.empty:
            print(f"\nMemproses sheet: {SHEET_NAME_PRODUCTION}")
            df_combined_production = df_production
            if excel_buffer is not None:
                excel_buffer.seek(0)
                try:
                    existing_df = pd.read_excel(excel_buffer, sheet_name=SHEET_NAME_PRODUCTION, engine='openpyxl')
                    df_combined_production = pd.concat([existing_df, df_production], ignore_index=True)
                    df_combined_production.drop_duplicates(subset=["Year"], keep="last", inplace=True)
                    df_combined_production = df_combined_production.sort_values("Year")
                    print(f"  Data di-merge. Total rows: {len(df_combined_production)}")
                except:
                    print(f"  Sheet baru akan dibuat")
            if SHEET_NAME_PRODUCTION in wb.sheetnames:
                del wb[SHEET_NAME_PRODUCTION]
            ws = wb.create_sheet(SHEET_NAME_PRODUCTION)
            for col_idx, col_name in enumerate(df_combined_production.columns, 1):
                ws.cell(row=1, column=col_idx, value=col_name)
            for row_idx, row_data in enumerate(df_combined_production.values, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            print(f"Sheet '{SHEET_NAME_PRODUCTION}': {len(df_combined_production)} rows")
        if df_country is not None and not df_country.empty:
            print(f"\nMemproses sheet: {SHEET_NAME_COUNTRY_STATS}")
            df_combined_country = df_country
            if excel_buffer is not None:
                excel_buffer.seek(0)
                try:
                    existing_df = pd.read_excel(excel_buffer, sheet_name=SHEET_NAME_COUNTRY_STATS, engine="openpyxl")
                    df_combined_country = pd.concat([existing_df, df_country], ignore_index=True)

                    # dedup paling aman by Country (atau by CountryCode kalau kamu yakin selalu terisi)
                    dedup_key = "CountryCode" if "CountryCode" in df_combined_country.columns else "Country"
                    df_combined_country.drop_duplicates(subset=[dedup_key], keep="last", inplace=True)

                    # sort optional
                    df_combined_country = df_combined_country.sort_values("Country")
                    print(f"  Data di-merge. Total rows: {len(df_combined_country)}")
                except:
                    print("  Sheet baru akan dibuat")

            if SHEET_NAME_COUNTRY_STATS in wb.sheetnames:
                del wb[SHEET_NAME_COUNTRY_STATS]
            ws = wb.create_sheet(SHEET_NAME_COUNTRY_STATS)

            for col_idx, col_name in enumerate(df_combined_country.columns, 1):
                ws.cell(row=1, column=col_idx, value=col_name)
            for row_idx, row_data in enumerate(df_combined_country.values, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

            print(f"Sheet '{SHEET_NAME_COUNTRY_STATS}': {len(df_combined_country)} rows")
        wb.save(output_buffer)
        wb.close()
        output_buffer.seek(0)
        verify_wb = load_workbook(output_buffer)
        print(f"\nVerifikasi - Sheets di buffer: {verify_wb.sheetnames}")
        verify_wb.close()
        output_buffer.seek(0)
        print(f"\nUploading ke OneDrive: {ONEDRIVE_FILE_PATH}")
        upload_excel_to_onedrive(access_token, ONEDRIVE_FILE_PATH, output_buffer)
        print("Upload selesai!")
    except Exception as e:
        print(f"Error saving to OneDrive: {e}")
        import traceback
        traceback.print_exc()

def main_iaea_scraper():
    try:
        access_token = get_access_token()
        print("Authentication successful\n")
    except Exception as e:
        print(f"Authentication failed: {e}")
        return
    df_capacity = fetch_nuclear_capacity_data()
    df_production = fetch_electrical_production_data()
    df_country = fetch_country_statistics_data()
    print(df_production)
    if (df_capacity is not None and not df_capacity.empty) or \
       (df_production is not None and not df_production.empty) or \
       (df_country is not None and not df_country.empty):
        save_to_onedrive(access_token, df_capacity, df_production,  df_country)
    else:
        print("\nTidak ada data yang berhasil diambil dari kedua halaman")
    print("SCRAPING SELESAI")

if __name__ == "__main__":
    main_iaea_scraper()