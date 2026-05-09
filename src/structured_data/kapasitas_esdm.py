import os
import sys
import traceback
from io import BytesIO

import pandas as pd
import requests
from dotenv import load_dotenv
from openpyxl import load_workbook

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from helpers.onedrive_helper import (
    download_excel_from_onedrive,
    get_access_token,
    upload_excel_to_onedrive,
)

load_dotenv()


# Constants

ONEDRIVE_FILE_PATH = os.getenv("ONEDRIVE_DATA_PATH", "/results/(Terstruktur)Data_Scraping_final.xlsx")
SHEET_NAME         = "(Data)Kapasitas_EBT"

API_URL     = "https://ebtke.esdm.go.id/api/api/konten/data-angka"
API_TIMEOUT = 30

MONTHS_EN_TO_NUM = {
    "January": 1,  "February": 2,  "March": 3,    "April": 4,
    "May": 5,      "June": 6,      "July": 7,      "August": 8,
    "September": 9,"October": 10,  "November": 11, "December": 12,
}

KAPASITAS_FIELDS = [
    "plta", "pltm", "pltmh", "pltp", "plts", "plts_atap",
    "pltb", "pltbm", "pltbg", "pltsa", "pltbn", "plt_hybrid", "total",
]


# API Fetch

def fetch_ebtke_data():
    """
    Fetch the latest EBT power plant capacity data from the EBTKE API.

    Returns a single-row DataFrame, or None on failure.
    """
    print("[Fetch] Mengambil data kapasitas pembangkit EBT...")
    try:
        response = requests.get(API_URL, timeout=API_TIMEOUT)
        response.raise_for_status()

        data          = response.json()
        kapasitas_data = data.get("data", {}).get("dataAngkaKapasitasPembangkit", {})

        if not kapasitas_data:
            print("[Fetch] Data kapasitas pembangkit tidak ditemukan di response API.")
            return None

        bulan = MONTHS_EN_TO_NUM.get(kapasitas_data.get("bulan"), kapasitas_data.get("bulan"))
        row   = {
            "tahun": kapasitas_data.get("tahun"),
            "bulan": bulan,
            **{field: kapasitas_data.get(field) for field in KAPASITAS_FIELDS},
        }

        df = pd.DataFrame([row])
        print(f"[Fetch] Data berhasil diambil: tahun={row['tahun']}, bulan={row['bulan']}.")
        return df

    except requests.exceptions.RequestException as exc:
        print(f"[Fetch] Error request ke API: {exc}")
        traceback.print_exc()
        return None
    except Exception as exc:
        print(f"[Fetch] Error tidak terduga: {exc}")
        traceback.print_exc()
        return None


# Duplicate Check

def check_data_exists(access_token, tahun, bulan):
    """
    Check if a row with the given tahun and bulan already exists in the OneDrive sheet.

    Returns True if found, False otherwise.
    """
    try:
        excel_buffer = download_excel_from_onedrive(access_token, ONEDRIVE_FILE_PATH)
        if excel_buffer is None:
            print("[Check] File belum ada di OneDrive.")
            return False

        excel_buffer.seek(0)
        wb = load_workbook(excel_buffer)

        if SHEET_NAME not in wb.sheetnames:
            print(f"[Check] Sheet '{SHEET_NAME}' belum ada.")
            wb.close()
            return False

        ws        = wb[SHEET_NAME]
        tahun_col = None
        bulan_col = None

        for col in range(1, ws.max_column + 1):
            header = ws.cell(1, col).value
            if header == "tahun":
                tahun_col = col
            elif header == "bulan":
                bulan_col = col

        if tahun_col is None or bulan_col is None:
            print("[Check] Kolom 'tahun' atau 'bulan' tidak ditemukan.")
            wb.close()
            return False

        for row in range(2, ws.max_row + 1):
            if ws.cell(row, tahun_col).value == tahun and ws.cell(row, bulan_col).value == bulan:
                print(f"[Check] Data tahun {tahun} bulan {bulan} sudah ada di OneDrive.")
                wb.close()
                return True

        print(f"[Check] Data tahun {tahun} bulan {bulan} belum ada.")
        wb.close()
        return False

    except Exception as exc:
        print(f"[Check] Error saat cek data: {exc}")
        traceback.print_exc()
        return False


# Data Merge

def _merge_data(df_existing, df_new):
    """
    Concatenate existing and new DataFrames, deduplicate by tahun+bulan, and sort.
    """
    df_combined = pd.concat([df_existing, df_new], ignore_index=True)

    if "tahun" in df_combined.columns and "bulan" in df_combined.columns:
        df_combined = df_combined.drop_duplicates(subset=["tahun", "bulan"], keep="last")

    if "tahun" in df_combined.columns:
        df_combined["tahun"] = pd.to_numeric(df_combined["tahun"], errors="coerce")
        df_combined = df_combined.sort_values(["tahun", "bulan"], ascending=True)

    print(f"[Merge] Data setelah merge: {len(df_combined)} rows.")
    return df_combined


# Save to OneDrive

def save_to_onedrive(access_token, df_new):
    """
    Merge new capacity data with existing OneDrive sheet, deduplicate, sort, and upload.

    Preserves all other sheets in the workbook.
    """
    if df_new is None or df_new.empty:
        print("[Save] Tidak ada data untuk disimpan.")
        return

    print(f"\n{'='*60}")
    print("[Save] Menyimpan data ke OneDrive")
    print(f"{'='*60}")

    excel_buffer  = download_excel_from_onedrive(access_token, ONEDRIVE_FILE_PATH)
    output_buffer = BytesIO()

    try:
        if excel_buffer is None:
            print("[Save] File belum ada — membuat baru...")
            with pd.ExcelWriter(output_buffer, engine="openpyxl", mode="w") as writer:
                df_new.to_excel(writer, sheet_name=SHEET_NAME, index=False)
            print(f"[Save] Sheet '{SHEET_NAME}' dibuat dengan {len(df_new)} rows.")
        else:
            print("[Save] File ditemukan — merging data...")
            excel_buffer.seek(0)
            wb = load_workbook(excel_buffer)

            # Fix hidden sheets
            visible_sheets = [s for s in wb.worksheets if s.sheet_state == "visible"]
            if len(visible_sheets) == 0:
                wb.worksheets[0].sheet_state = "visible"
                wb.active = 0

            if SHEET_NAME in wb.sheetnames:
                ws      = wb[SHEET_NAME]
                headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
                rows    = [
                    [ws.cell(row, col).value for col in range(1, ws.max_column + 1)]
                    for row in range(2, ws.max_row + 1)
                ]
                df_existing = pd.DataFrame(rows, columns=headers)
                print(f"[Save] Data existing : {len(df_existing)} rows.")
                print(f"[Save] Data baru     : {len(df_new)} rows.")
                df_merged = _merge_data(df_existing, df_new)
                del wb[SHEET_NAME]
            else:
                print(f"[Save] Sheet '{SHEET_NAME}' belum ada — membuat baru...")
                df_merged = df_new

            ws = wb.create_sheet(SHEET_NAME)
            for col_idx, col_name in enumerate(df_merged.columns, 1):
                ws.cell(row=1, column=col_idx, value=col_name)
            for row_idx, row_data in enumerate(df_merged.values, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

            print(f"[Save] Total setelah merge: {len(df_merged)} rows.")
            wb.save(output_buffer)
            wb.close()

        output_buffer.seek(0)

        # Verifikasi
        verify_wb = load_workbook(output_buffer)
        print(f"[Save] Verifikasi sheet: {verify_wb.sheetnames}")
        verify_wb.close()
        output_buffer.seek(0)

        # Upload
        print(f"[Save] Uploading ke OneDrive: {ONEDRIVE_FILE_PATH}")
        upload_excel_to_onedrive(access_token, ONEDRIVE_FILE_PATH, output_buffer)

        print(f"\n{'='*60}")
        print("[Save] DATA BERHASIL DISIMPAN KE ONEDRIVE")
        print(f"{'='*60}")
        print(f"[Save] File : {ONEDRIVE_FILE_PATH}")
        print(f"[Save] Sheet: {SHEET_NAME}")

    except Exception as exc:
        print(f"[Save] Error saat menyimpan data: {exc}")
        traceback.print_exc()


# Public Entry Point

def main_ebtke_scraper():
    """
    Run the full EBT capacity scraping workflow:
    authenticate, fetch latest data, check duplicate, save to OneDrive.
    """
    print(f"\n{'='*60}")
    print("SCRAPER KAPASITAS EBT (EBTKE API)")
    print("STORAGE MODE: OneDrive")
    print(f"{'='*60}")
    print(f"\n[Main] File : {ONEDRIVE_FILE_PATH}")
    print(f"[Main] Sheet: {SHEET_NAME}")

    print("\n[Main] Authenticating to Microsoft Graph API...")
    try:
        access_token = get_access_token()
        print("[Main] Authentication successful.")
    except Exception as exc:
        print(f"[Main] Authentication failed: {exc}")
        return

    df_new = fetch_ebtke_data()
    if df_new is None or df_new.empty:
        print("\n[Main] Tidak ada data yang berhasil diambil.")
        return

    tahun = df_new.iloc[0]["tahun"]
    bulan = df_new.iloc[0]["bulan"]

    if check_data_exists(access_token, tahun, bulan):
        print(f"\n[Main] Data tahun {tahun} bulan {bulan} sudah ada — skip upload.")
        return

    save_to_onedrive(access_token, df_new)

    print(f"\n{'='*60}")
    print("[Main] SELESAI!")
    print(f"{'='*60}\n")


# Script Entry Point

if __name__ == "__main__":
    main_ebtke_scraper()