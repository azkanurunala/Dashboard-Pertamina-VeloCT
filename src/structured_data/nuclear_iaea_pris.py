import os
import re
import sys
import time
import traceback
from datetime import date, datetime
from io import BytesIO
from urllib.parse import parse_qs, urlparse

import pandas as pd
import requests
from bs4 import BeautifulSoup
from dotenv import load_dotenv
from openpyxl import load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from helpers.onedrive_helper import (
    download_excel_from_onedrive,
    get_access_token,
    upload_excel_to_onedrive,
)
from helpers.scraping_helper import setup_driver

load_dotenv()


# Constants

ONEDRIVE_FILE_PATH = os.getenv("ONEDRIVE_DATA_PATH", "/results/(Terstruktur)Data_Scraping_final.xlsx")

SHEET_NAME_CAPACITY     = "(Data)IAEA_Nuclear_Capacity"
SHEET_NAME_PRODUCTION   = "(Data)IAEA_Electrical"
SHEET_NAME_COUNTRY_STATS = "(Data)IAEA_Country_Stats"

URL_CAPACITY     = "https://pris.iaea.org/PRIS/WorldStatistics/WorldTrendNuclearPowerCapacity.aspx"
URL_PRODUCTION   = "https://pris.iaea.org/PRIS/WorldStatistics/WorldTrendinElectricalProduction.aspx"
URL_COUNTRY_STATS = "https://pris.iaea.org/PRIS/CountryStatistics/CountryStatisticsLandingPage.aspx"
URL_LAST_UPDATE  = "https://pris.iaea.org/PRIS/WorldStatistics/OperationalReactorsByCountry.aspx"

SELENIUM_WAIT    = 20
SELENIUM_SLEEP   = 3


# Parsing Utilities

def _to_number(x: str):
    """Convert a string cell value to int, float, or return as-is."""
    x = (x or "").strip()
    if x == "":
        return None
    try:
        return int(x)
    except ValueError:
        try:
            return float(x)
        except ValueError:
            return x

def _parse_hash_pipe_line(line: str, col_names: list[str]) -> pd.DataFrame:
    """
    Parse a stats line of format 'COUNTRY|a|b|c#COUNTRY2|a|b|c...' into a DataFrame.
    """
    records = []
    for item in line.split("#"):
        parts = [p.strip() for p in item.split("|")]
        if len(parts) != len(col_names):
            continue
        rec = {"Country": parts[0].strip()}
        for i, c in enumerate(col_names[1:], start=1):
            rec[c] = _to_number(parts[i])
        records.append(rec)
    return pd.DataFrame(records)

def _parse_table_from_soup(soup: BeautifulSoup, skip_empty_cells: bool = False) -> tuple[list, list]:
    """
    Extract headers and rows from the first <table> found in BeautifulSoup object.

    Returns (headers, data) lists.
    """
    table = soup.find("table")
    if not table:
        return [], []

    headers = []
    thead   = table.find("thead")
    if thead:
        for th in thead.find_all("th"):
            text = th.get_text(strip=True).replace("\n", " ")
            if not skip_empty_cells or text:
                headers.append(text)

    data  = []
    tbody = table.find("tbody")
    if tbody:
        for tr in tbody.find_all("tr"):
            row = []
            for td in tr.find_all("td"):
                text = td.get_text(strip=True)
                if skip_empty_cells and (not text or text == "\xa0"):
                    continue
                try:
                    row.append(float(text) if "." in text else int(text))
                except ValueError:
                    row.append(text)
            if row and (not skip_empty_cells or len(row) == len(headers)):
                data.append(row)

    return headers, data


# Data Fetching

def fetch_nuclear_capacity_data() -> pd.DataFrame | None:
    """
    Scrape world nuclear power capacity trend table from IAEA PRIS.

    Returns a DataFrame or None on failure.
    """
    driver = None
    try:
        driver = setup_driver()
        driver.get(URL_CAPACITY)
        WebDriverWait(driver, SELENIUM_WAIT).until(
            EC.presence_of_element_located((By.TAG_NAME, "table"))
        )
        time.sleep(SELENIUM_SLEEP)

        soup    = BeautifulSoup(driver.page_source, "html.parser")
        headers, data = _parse_table_from_soup(soup)

        if not headers or not data:
            print("[Fetch] Tabel Nuclear Capacity tidak ditemukan.")
            return None

        df = pd.DataFrame(data, columns=headers)
        print(f"[Fetch] Nuclear Capacity: {len(df)} baris.")
        return df

    except Exception as exc:
        print(f"[Fetch] Error Nuclear Capacity: {exc}")
        return None
    finally:
        if driver:
            driver.quit()

def fetch_electrical_production_data() -> pd.DataFrame | None:
    """
    Scrape world electrical production trend table from IAEA PRIS.

    Returns a DataFrame or None on failure.
    """
    driver = None
    try:
        driver = setup_driver()
        driver.get(URL_PRODUCTION)
        WebDriverWait(driver, SELENIUM_WAIT).until(
            EC.presence_of_element_located((By.TAG_NAME, "table"))
        )
        time.sleep(SELENIUM_SLEEP)

        soup    = BeautifulSoup(driver.page_source, "html.parser")
        headers, data = _parse_table_from_soup(soup, skip_empty_cells=True)

        if not headers or not data:
            print("[Fetch] Tabel Electrical Production tidak ditemukan.")
            return None

        df = pd.DataFrame(data, columns=headers)
        print(f"[Fetch] Electrical Production: {len(df)} baris.")
        return df

    except Exception as exc:
        print(f"[Fetch] Error Electrical Production: {exc}")
        traceback.print_exc()
        return None
    finally:
        if driver:
            driver.quit()

def _fetch_pris_last_update() -> date | None:
    """Scrape the PRIS last update date from the operational reactors page."""
    driver = None
    try:
        driver = setup_driver()
        driver.get(URL_LAST_UPDATE)
        WebDriverWait(driver, SELENIUM_WAIT).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )
        hits = re.findall(r"Last update on\s+(\d{4}-\d{2}-\d{2})", driver.page_source)
        return datetime.strptime(hits[-1], "%Y-%m-%d").date() if hits else None
    finally:
        if driver:
            driver.quit()

def fetch_country_statistics_data() -> pd.DataFrame | None:
    """
    Scrape country-level nuclear statistics from IAEA PRIS landing page.

    Merges reactor counts, capacity, and country codes into a single DataFrame.
    Adds a LastUpdate column from the PRIS operational reactors page.
    Returns a DataFrame or None on failure.
    """
    driver = None
    try:
        driver = setup_driver()
        driver.get(URL_COUNTRY_STATS)
        WebDriverWait(driver, SELENIUM_WAIT).until(
            EC.presence_of_element_located((By.TAG_NAME, "body"))
        )
        time.sleep(2)

        soup  = BeautifulSoup(driver.page_source, "html.parser")
        lines = [l.strip() for l in soup.get_text("\n").splitlines() if l.strip()]
        stats_lines = [l for l in lines if "#" in l and "|" in l]

        if len(stats_lines) < 2:
            print("[Fetch] Tidak menemukan 2 baris statistik country.")
            return None

        df_reactors = _parse_hash_pipe_line(
            stats_lines[0],
            ["Country", "Reactors_InOperation", "Reactors_UnderConstruction", "Reactors_PermanentShutdown"],
        )
        df_capacity = _parse_hash_pipe_line(
            stats_lines[1],
            ["Country", "NetCapacityMW_InOperation", "NetCapacityMW_UnderConstruction", "NetCapacityMW_PermanentShutdown"],
        )

        # Build country code map from href params
        code_map = {}
        for a in soup.select("a[href*='CountryDetails.aspx?current=']"):
            name = (a.get_text(strip=True) or "").strip()
            qs   = parse_qs(urlparse(a.get("href", "")).query)
            cc   = (qs.get("current", [None])[0] or "").strip()
            if name and cc:
                code_map[name.upper()] = cc

        df = df_reactors.merge(df_capacity, on="Country", how="outer")
        df["CountryCode"] = df["Country"].astype(str).str.upper().map(code_map)

        pris_last = _fetch_pris_last_update()
        if pris_last is None:
            raise ValueError("Tidak menemukan 'Last update on YYYY-MM-DD' di PRIS.")

        df["LastUpdate"] = pris_last
        cols = ["LastUpdate", "Country", "CountryCode"] + [
            c for c in df.columns if c not in ("LastUpdate", "Country", "CountryCode")
        ]
        df = df[cols]

        print(f"[Fetch] Country Statistics: {len(df)} baris, LastUpdate={pris_last}.")
        return df

    except Exception as exc:
        print(f"[Fetch] Error Country Statistics: {exc}")
        traceback.print_exc()
        return None
    finally:
        if driver:
            driver.quit()


# Save to OneDrive

def _write_sheet(wb, sheet_name: str, df: pd.DataFrame) -> None:
    """Delete and recreate a sheet in the workbook with DataFrame content."""
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)
    for col_idx, col_name in enumerate(df.columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)
    for row_idx, row_data in enumerate(df.values, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

def save_to_onedrive(
    access_token,
    df_capacity: pd.DataFrame | None,
    df_production: pd.DataFrame | None,
    df_country: pd.DataFrame | None,
) -> None:
    """
    Merge and upload all three IAEA DataFrames to OneDrive.

    Preserves all other sheets in the workbook.
    Each sheet is deduplicated and sorted before saving.
    """
    if all(df is None or df.empty for df in [df_capacity, df_production, df_country]):
        print("[Save] Tidak ada data untuk disimpan.")
        return

    print(f"\n{'='*60}")
    print("[Save] Menyimpan data ke OneDrive")
    print(f"{'='*60}")

    excel_buffer  = download_excel_from_onedrive(access_token, ONEDRIVE_FILE_PATH)
    output_buffer = BytesIO()

    try:
        if excel_buffer is None:
            print("[Save] File tidak ada di OneDrive — membuat baru...")
            wb = load_workbook(BytesIO())
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]
        else:
            print("[Save] File ditemukan di OneDrive — updating...")
            excel_buffer.seek(0)
            wb = load_workbook(excel_buffer)

            # Fix hidden sheets
            if not any(s.sheet_state == "visible" for s in wb.worksheets):
                wb.worksheets[0].sheet_state = "visible"
                wb.active = 0

        # --- Nuclear Capacity ---
        if df_capacity is not None and not df_capacity.empty:
            print(f"\n[Save] Sheet: {SHEET_NAME_CAPACITY}")
            df_combined = df_capacity
            if excel_buffer is not None:
                try:
                    excel_buffer.seek(0)
                    existing = pd.read_excel(excel_buffer, sheet_name=SHEET_NAME_CAPACITY, engine="openpyxl")
                    df_combined = pd.concat([existing, df_capacity], ignore_index=True)
                    df_combined.drop_duplicates(subset=["Year"], keep="last", inplace=True)
                    df_combined.sort_values("Year", inplace=True)
                    print(f"[Save] Merged: {len(df_combined)} rows.")
                except Exception:
                    print("[Save] Sheet baru akan dibuat.")
            _write_sheet(wb, SHEET_NAME_CAPACITY, df_combined)
            print(f"[Save] {SHEET_NAME_CAPACITY}: {len(df_combined)} rows.")

        # --- Electrical Production ---
        if df_production is not None and not df_production.empty:
            print(f"\n[Save] Sheet: {SHEET_NAME_PRODUCTION}")
            df_combined = df_production
            if excel_buffer is not None:
                try:
                    excel_buffer.seek(0)
                    existing = pd.read_excel(excel_buffer, sheet_name=SHEET_NAME_PRODUCTION, engine="openpyxl")
                    df_combined = pd.concat([existing, df_production], ignore_index=True)
                    df_combined.drop_duplicates(subset=["Year"], keep="last", inplace=True)
                    df_combined.sort_values("Year", inplace=True)
                    print(f"[Save] Merged: {len(df_combined)} rows.")
                except Exception:
                    print("[Save] Sheet baru akan dibuat.")
            _write_sheet(wb, SHEET_NAME_PRODUCTION, df_combined)
            print(f"[Save] {SHEET_NAME_PRODUCTION}: {len(df_combined)} rows.")

        # --- Country Statistics ---
        if df_country is not None and not df_country.empty:
            print(f"\n[Save] Sheet: {SHEET_NAME_COUNTRY_STATS}")
            df_combined = df_country
            if excel_buffer is not None:
                try:
                    excel_buffer.seek(0)
                    existing = pd.read_excel(excel_buffer, sheet_name=SHEET_NAME_COUNTRY_STATS, engine="openpyxl")
                    pris_last = df_country["LastUpdate"].iloc[0]

                    if "LastUpdate" in existing.columns:
                        last_saved = pd.to_datetime(existing["LastUpdate"], errors="coerce").max()
                        last_saved = last_saved.date() if pd.notna(last_saved) else None
                        if last_saved == pris_last:
                            print(f"[Save] LastUpdate sama ({pris_last}) — skip append.")
                            df_combined = existing
                        else:
                            df_combined = pd.concat([existing, df_country], ignore_index=True)
                    else:
                        df_combined = pd.concat([existing, df_country], ignore_index=True)

                except Exception:
                    print("[Save] Sheet baru akan dibuat.")

            key = "CountryCode" if "CountryCode" in df_combined.columns else "Country"
            df_combined.drop_duplicates(subset=["LastUpdate", key], keep="last", inplace=True)
            df_combined.sort_values(["LastUpdate", "Country"], ascending=True, inplace=True)
            _write_sheet(wb, SHEET_NAME_COUNTRY_STATS, df_combined)
            print(f"[Save] {SHEET_NAME_COUNTRY_STATS}: {len(df_combined)} rows.")

        wb.save(output_buffer)
        wb.close()
        output_buffer.seek(0)

        # Verifikasi
        verify_wb = load_workbook(output_buffer)
        print(f"\n[Save] Verifikasi sheet: {verify_wb.sheetnames}")
        verify_wb.close()
        output_buffer.seek(0)

        # Upload
        print(f"[Save] Uploading ke OneDrive: {ONEDRIVE_FILE_PATH}")
        upload_excel_to_onedrive(access_token, ONEDRIVE_FILE_PATH, output_buffer)

        print(f"\n{'='*60}")
        print("[Save] DATA BERHASIL DISIMPAN KE ONEDRIVE")
        print(f"{'='*60}")
        print(f"[Save] File: {ONEDRIVE_FILE_PATH}")

    except Exception as exc:
        print(f"[Save] Error: {exc}")
        traceback.print_exc()


# Public Entry Point

def main_iaea_scraper() -> None:
    """
    Run the full IAEA PRIS scraping workflow:
    authenticate, scrape capacity/production/country data, save to OneDrive.
    """
    print(f"\n{'='*60}")
    print("SCRAPER IAEA PRIS")
    print("STORAGE MODE: OneDrive")
    print(f"{'='*60}")
    print(f"\n[Main] File: {ONEDRIVE_FILE_PATH}")

    print("\n[Main] Authenticating to Microsoft Graph API...")
    try:
        access_token = get_access_token()
        print("[Main] Authentication successful.")
    except Exception as exc:
        print(f"[Main] Authentication failed: {exc}")
        return

    print("\n[Main] Fetching Nuclear Capacity...")
    df_capacity = fetch_nuclear_capacity_data()

    print("\n[Main] Fetching Electrical Production...")
    df_production = fetch_electrical_production_data()

    print("\n[Main] Fetching Country Statistics...")
    df_country = fetch_country_statistics_data()

    if all(df is None or df.empty for df in [df_capacity, df_production, df_country]):
        print("\n[Main] Tidak ada data yang berhasil diambil.")
        return

    save_to_onedrive(access_token, df_capacity, df_production, df_country)

    print(f"\n{'='*60}")
    print("[Main] SELESAI!")
    print(f"{'='*60}\n")


# Script Entry Point

if __name__ == "__main__":
    main_iaea_scraper()