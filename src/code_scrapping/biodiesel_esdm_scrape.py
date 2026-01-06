import requests
import pandas as pd
import time
import re
import os
import pdfplumber
from datetime import datetime
from tqdm import tqdm
import sys
from dotenv import load_dotenv
from io import BytesIO
from openpyxl import load_workbook

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from helpers.onedrive_helper import (
    get_access_token,
    download_excel_from_onedrive,
    upload_excel_to_onedrive
)

load_dotenv()

ONEDRIVE_FILE_PATH = os.getenv("ONEDRIVE_DATA_PATH", "/results/(Terstruktur)Data Scrapping.xlsx")
SHEET_NAME = "(Data)Biodesel"

def parse_date(date_str):
    months = {
        'Januari': '01', 'Februari': '02', 'Maret': '03', 'April': '04',
        'Mei': '05', 'Juni': '06', 'Juli': '07', 'Agustus': '08',
        'September': '09', 'Oktober': '10', 'November': '11', 'Desember': '12'
    }
    date_match = re.search(r'(\d{1,2})\s+(\w+)\s+(\d{4})', date_str)
    if date_match:
        day = date_match.group(1).zfill(2)
        month = months.get(date_match.group(2))
        year = date_match.group(3)
        return f"{year}-{month}-{day}"
    return None

def matches_biodiesel_criteria(title):
    keywords = ["HIP", "BBN", "JENIS", "BIODIESEL", "BULAN"]
    return all(keyword in title.upper() for keyword in keywords)

def get_missing_months_from_excel(access_token, sheet_name='(Data)Biodesel'):
    try:
        excel_buffer = download_excel_from_onedrive(access_token, ONEDRIVE_FILE_PATH)
        if excel_buffer is None:
            print("File Excel tidak ditemukan di OneDrive, asumsikan scraping awal")
            return None
        df = pd.read_excel(excel_buffer, sheet_name=sheet_name, engine='openpyxl')  
        if df.empty or 'Bulan HIP' not in df.columns:
            print("Sheet kosong atau kolom 'Bulan HIP' tidak ada, asumsikan scraping awal")
            return None
        df = df[df['Bulan HIP'].notna()]
        if df.empty:
            print("Tidak ada data Bulan HIP, asumsikan scraping awal")
            return None
        months_map = {
            'Januari': 1, 'Februari': 2, 'Maret': 3, 'April': 4,
            'Mei': 5, 'Juni': 6, 'Juli': 7, 'Agustus': 8,
            'September': 9, 'Oktober': 10, 'November': 11, 'Desember': 12
        }
        def parse_bulan_hip(bulan_str):
            try:
                parts = bulan_str.strip().split()
                if len(parts) >= 2:
                    month_name = parts[0]
                    year = int(parts[-1])  
                    month = months_map.get(month_name)
                    if month and year:
                        return pd.Timestamp(year=year, month=month, day=1)
            except:
                pass
            return None
        df['parsed_bulan_hip'] = df['Bulan HIP'].apply(parse_bulan_hip)
        df = df[df['parsed_bulan_hip'].notna()]
        if df.empty:
            print("Tidak bisa parse Bulan HIP, asumsikan scraping awal")
            return None
        last_date = df['parsed_bulan_hip'].max()
        last_month = last_date.month
        last_year = last_date.year
        now = datetime.now()
        current_month = now.month
        current_year = now.year
        diff = (current_year - last_year) * 12 + (current_month - last_month)
        
        if diff <= 0:
            print(f"Data sudah up-to-date (bulan terakhir: {last_month}/{last_year})")
            return 0
        month_names_id = {
            1: 'Januari', 2: 'Februari', 3: 'Maret', 4: 'April',
            5: 'Mei', 6: 'Juni', 7: 'Juli', 8: 'Agustus',
            9: 'September', 10: 'Oktober', 11: 'November', 12: 'Desember'
        }
        last_bulan_hip = df.loc[df['parsed_bulan_hip'] == last_date, 'Bulan HIP'].iloc[0]
        print(f"Bulan HIP terakhir di file: {last_bulan_hip}")
        print(f"Bulan sekarang: {month_names_id[current_month]} {current_year}")
        print(f"Selisih: {diff} bulan")
        return diff
    except FileNotFoundError:
        print("File Excel tidak ditemukan, asumsikan scraping awal")
        return None
    except Exception as e:
        print(f"Error membaca file: {e}")
        print("Asumsikan scraping awal")
        return None

def extract_pdf_url_from_html(html_content):
    if not html_content:
        return None
    match = re.search(r'href=["\']([^"\']*drive\.esdm\.go\.id[^"\']*)["\']', html_content)
    if match:
        return match.group(1)
    return None

def scrape_biodiesel_articles_api(excel_filename='../results/(Terstruktur)Data Scrapping.xlsx', sheet_name='(Data)Biodesel'):
    missing_months = get_missing_months_from_excel(excel_filename, sheet_name)
    if missing_months == 0:
        print("Tidak ada artikel baru yang perlu diambil")
        return [], missing_months
    elif missing_months is None:
        print("File kosong, ambil semua artikel")
        length = 200 
    else:
        max_articles_per_month = 10
        length = missing_months * max_articles_per_month
        print(f"Target: {length} artikel ({missing_months} bulan)")
    base_api_url = "https://ebtke.esdm.go.id/api/api/artikel"
    api_url = f"{base_api_url}?kategori_slug=pengumuman&start=0&length={length}&is_published=true"
    print(f"\nMengambil {length} artikel dari API...")
    try:
        response = requests.get(api_url, timeout=30)
        response.raise_for_status()
        json_data = response.json()
        if 'data' not in json_data or not json_data['data']:
            print("Tidak ada data")
            return [], missing_months
        articles = json_data['data']
        print(f"Berhasil ambil {len(articles)} artikel dari API")
        data = []
        for article in articles:
            title = article.get('judul', '').strip()
            if matches_biodiesel_criteria(title):
                date_str = article.get('tanggal_publikasi') or article.get('tgl_upload', '')
                article_date = date_str
                slug = article.get('slug', '')
                article_url = f"https://ebtke.esdm.go.id/artikel/pengumuman/{slug}"
                konten = article.get('konten', '')
                pdf_url = extract_pdf_url_from_html(konten)
                data.append({
                    "Judul": title,
                    "url": article_url,
                    "Date": article_date,
                    "konten": konten,
                    "pdf_url": pdf_url
                })
        print(f"{len(data)} artikel biodiesel ditemukan")
        unique_data = []
        seen_keys = set()
        for item in data:
            key = (item.get("Judul"), item.get("Date"))
            if key not in seen_keys:
                seen_keys.add(key)
                unique_data.append(item)
        if len(unique_data) < len(data):
            print(f"Setelah deduplikasi: {len(unique_data)} artikel")
        return unique_data, missing_months
    except requests.exceptions.RequestException as e:
        print(f"Error API: {e}")
        return [], missing_months

def download_pdf(url, filename):
    try:
        if not url.startswith('http'):
            url = 'https://' + url
        if 'drive.esdm.go.id' in url and 'download' not in url:
            url = url + '&mode=list&download=1' if '?' in url else url + '?download=1'
        response = requests.get(url, stream=True, timeout=30)
        if 'application/pdf' not in response.headers.get('content-type', '').lower():
            return False
        with open(filename, 'wb') as f:
            for chunk in response.iter_content(chunk_size=8192):
                if chunk:
                    f.write(chunk)
        return True
    except Exception as e:
        print(f"Error download: {e}")
        return False

def scrape_and_download_pdfs(data, missing_months):
    pdf_links = []
    if missing_months == 0:
        print("Tidak ada artikel baru, semua data sudah lengkap.")
        return pdf_links
    elif missing_months is None:
        print("Scraping awal: ambil semua artikel yang tersedia.")
        filtered_data = data
    else:
        filtered_data = data[:missing_months]
        print(f"Menargetkan {len(filtered_data)} artikel terbaru (untuk {missing_months} bulan hilang).")
    for item in tqdm(filtered_data, desc="Mencari file PDF"):
        pdf_url = item.get('pdf_url')
        if not pdf_url:
            print(f"\nTidak ada PDF URL: {item['Judul'][:50]}...")
            continue
        filename = f"HIP_BBN_{item['Date']}.pdf".replace(":", "-")
        if download_pdf(pdf_url, filename):
            item["pdf_filename"] = filename
            pdf_links.append(item)
        else:
            print(f"\nGagal mendownload PDF dari {pdf_url}")
        time.sleep(0.3)
    return pdf_links

def find_hip_value_and_month_in_table(table):
    hip_value = None
    hip_month = None
    for row_idx, row in enumerate(table):
        if not row:
            continue
        for col_idx, cell in enumerate(row):
            text = str(cell) if cell else ""
            if '(RUPIAH/LITER)' in text.upper():
                if row_idx + 1 >= len(table):
                    continue
                next_row = table[row_idx + 1]
                for val in reversed(next_row):
                    if val:
                        val_clean = str(val).replace(',', '.').replace(' ', '').strip()
                        match = re.match(r'^(\d+(?:\.\d+)?)$', val_clean)
                        if match:
                            hip_value = float(match.group(1))
                            break
                for val in reversed(next_row):
                    if isinstance(val, str) and re.search(
                        r'(Januari|Februari|Maret|April|Mei|Juni|Juli|Agustus|September|Oktober|November|Desember)\s+\d{4}',
                        val
                    ):
                        hip_month = val.strip()
                        break
                if hip_value:
                    return hip_value, hip_month
    return hip_value, hip_month

def extract_hip_from_pdf(pdf_file):
    try:
        with pdfplumber.open(pdf_file) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables()
                for table in tables:
                    hip_value, hip_month = find_hip_value_and_month_in_table(table)
                    if hip_value:
                        return hip_value, hip_month
        return None, None
    except Exception as e:
        print(f"Error parsing {pdf_file}: {e}")
        return None, None

def parse_all_pdfs(pdf_links):
    excel_data = []
    for item in pdf_links:
        pdf_file = item.get('pdf_filename')
        if not pdf_file or not os.path.exists(pdf_file):
            print(f"PDF tidak ditemukan: {pdf_file}")
            continue
        print(f"\nParsing {pdf_file}")
        hip_per_liter, hip_month = extract_hip_from_pdf(pdf_file)
        date_artikel = item.get('Date', None)
        if hip_per_liter:
            print(f"HIP Biodiesel IDR/L: {hip_per_liter}")
            if hip_month:
                print(f"Bulan HIP: {hip_month}")
            print(f"Date artikel: {date_artikel}")
            excel_data.append({
                'Date': date_artikel,
                'Bulan HIP': hip_month,
                'HIP Biodiesel IDR/L': hip_per_liter
            })
            try:
                os.remove(pdf_file)
                print(f"PDF dihapus: {pdf_file}")
            except Exception as e:
                print(f"Gagal menghapus {pdf_file}: {e}")
    return excel_data

def save_to_excel(access_token, data, sheet_name='(Data)Biodesel'):
    if not data:
        print("Tidak ada data baru untuk disimpan")
        return None
    
    print("\n" + "="*60)
    print("Menyimpan data ke OneDrive")
    print("="*60)
    
    new_df = pd.DataFrame(data)
    new_df['HIP Biodiesel IDR/L'] = (
        new_df['HIP Biodiesel IDR/L']
        .astype(float)
        .mul(1000)
        .astype(int)
    )
    new_df['Date'] = pd.to_datetime(new_df['Date'], errors='coerce')
    
    print(f"Data baru: {len(new_df)} baris")
    
    # Baca data existing dari OneDrive
    excel_buffer = download_excel_from_onedrive(access_token, ONEDRIVE_FILE_PATH)
    
    if excel_buffer is None:
        print("File tidak ada di OneDrive, akan membuat baru")
        combined_df = new_df
    else:
        try:
            existing_df = pd.read_excel(excel_buffer, sheet_name=sheet_name, engine='openpyxl')
            existing_df['Date'] = pd.to_datetime(existing_df['Date'], errors='coerce')
            print(f"Data lama: {len(existing_df)} baris")
            combined_df = pd.concat([existing_df, new_df], ignore_index=True)
        except ValueError:
            print(f"Sheet '{sheet_name}' tidak ditemukan, membuat sheet baru")
            combined_df = new_df
        except Exception as e:
            print(f"Error membaca sheet existing: {e}")
            combined_df = new_df
    
    # Deduplikasi dan sorting
    combined_df = combined_df.drop_duplicates(subset=['Bulan HIP'], keep='last')
    combined_df = combined_df.sort_values(by='Date', ascending=True)
    combined_df['Date'] = combined_df['Date'].dt.strftime('%Y-%m-%d')
    
    print(f"Data setelah deduplikasi: {len(combined_df)} baris")
    
    # Siapkan buffer untuk upload
    from io import BytesIO
    from openpyxl import load_workbook, Workbook
    
    output_buffer = BytesIO()
    
    try:
        if excel_buffer is None:
            # File benar-benar tidak ada, buat baru
            print("Membuat file Excel baru...")
            with pd.ExcelWriter(output_buffer, engine='openpyxl', mode='w') as writer:
                combined_df.to_excel(writer, sheet_name=sheet_name, index=False)
        else:
            # File sudah ada - PRESERVE semua sheet
            print("File existing ditemukan, preserve semua sheet...")
            
            # Load workbook existing
            excel_buffer.seek(0)  # Reset pointer
            wb = load_workbook(excel_buffer)
            
            print(f"Sheet yang ada saat ini: {wb.sheetnames}")
            
            # Fix hidden sheets
            visible_sheets = [s for s in wb.worksheets if s.sheet_state == 'visible']
            if len(visible_sheets) == 0:
                print("Fixing hidden sheets...")
                wb.worksheets[0].sheet_state = 'visible'
                wb.active = 0
            
            for sheet in wb.worksheets:
                if sheet.sheet_state != 'visible':
                    sheet.sheet_state = 'visible'
            
            # Hapus sheet target jika sudah ada (akan dibuat ulang)
            if sheet_name in wb.sheetnames:
                print(f"Menghapus sheet '{sheet_name}' yang lama...")
                del wb[sheet_name]
            
            # Buat sheet baru dengan data
            ws = wb.create_sheet(sheet_name)
            
            # Tulis header
            for col_idx, col_name in enumerate(combined_df.columns, 1):
                ws.cell(row=1, column=col_idx, value=col_name)
            
            # Tulis data
            for row_idx, row_data in enumerate(combined_df.values, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            print(f"Sheet yang akan disimpan: {wb.sheetnames}")
            
            # Save ke buffer
            wb.save(output_buffer)
            wb.close()
        
        output_buffer.seek(0)
        
        # Verifikasi sheet sebelum upload
        verify_wb = load_workbook(output_buffer)
        print(f"Verifikasi - Sheet di buffer: {verify_wb.sheetnames}")
        verify_wb.close()
        output_buffer.seek(0)
        
        # Upload ke OneDrive
        print(f"\nUploading ke OneDrive: {ONEDRIVE_FILE_PATH}")
        upload_excel_to_onedrive(access_token, ONEDRIVE_FILE_PATH, output_buffer)
        print("Upload selesai!")
        
        print("\n" + "="*60)
        print("DATA BERHASIL DISIMPAN KE ONEDRIVE")
        print("="*60)
        print(f"  File: {ONEDRIVE_FILE_PATH}")
        print(f"  Sheet: {sheet_name}")
        print(f"  Total rows: {len(combined_df)} (sorted ascending by Date)")
        print(f"  Data baru ditambahkan: {len(new_df)} baris")
        
        return combined_df
        
    except Exception as e:
        print(f"Error saat menyimpan: {e}")
        import traceback
        traceback.print_exc()
        return None   

def main_biodiesel_esdm():
    print("="*60)
    print("SCRAPER HIP BBN BIODIESEL (API VERSION)")
    print("STORAGE MODE: OneDrive")
    print("="*60)
    print(f"\nFile: {ONEDRIVE_FILE_PATH}")
    print(f"Sheet: {SHEET_NAME}")
    print("\nAuthenticating to Microsoft Graph API...")
    try:
        access_token = get_access_token()
        print("Authentication successful")
    except Exception as e:
        print(f"Authentication failed: {e}")
        return
    data, missing_months = scrape_biodiesel_articles_api(access_token, SHEET_NAME)
    if not data:
        print("\nTidak ada data untuk diproses")
    else:
        print(f"\nTotal artikel biodiesel: {len(data)}")
        pdf_links = scrape_and_download_pdfs(data, missing_months)
        if not pdf_links:
            print("\nTidak ada PDF yang berhasil didownload")
        else:
            excel_data = parse_all_pdfs(pdf_links)
            if not excel_data:
                print("\nTidak ada data HIP yang berhasil di-extract")
            else:
                df = save_to_excel(access_token, excel_data, SHEET_NAME)
                if df is not None:
                    print("\n" + "="*60)
                    print("SELESAI!")
                    print("="*60)
                else:
                    print("\nGagal menyimpan data ke OneDrive")


if __name__ == "__main__":
    main_biodiesel_esdm()