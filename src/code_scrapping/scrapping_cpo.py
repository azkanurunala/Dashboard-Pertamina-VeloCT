import re
import requests
import pandas as pd
from bs4 import BeautifulSoup
from datetime import datetime
import os
import sys
from dotenv import load_dotenv

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from helpers.onedrive_helper import (
    get_access_token,
    download_excel_from_onedrive,
    upload_excel_to_onedrive
)

load_dotenv()

ONEDRIVE_FILE_PATH = os.getenv("ONEDRIVE_DATA_PATH", "/results/(Terstruktur)Data Scrapping.xlsx")
SHEET_NAME = "(Data)CPO"


def read_cpo_sheet_from_onedrive(access_token, file_path, sheet_name):
    excel_buffer = download_excel_from_onedrive(access_token, file_path)
    
    if excel_buffer is None:
        print(f"! File tidak ditemukan, akan membuat baru")
        return pd.DataFrame(columns=["Upload_Dates", "Dates", "PX_LAST"])
    
    try:
        df = pd.read_excel(excel_buffer, sheet_name=sheet_name)
        print(f"Berhasil baca sheet '{sheet_name}', rows={len(df)}")
        return df
    except Exception as e:
        print(f"! Sheet '{sheet_name}' tidak ditemukan, akan membuat baru")
        return pd.DataFrame(columns=["Upload_Dates", "Dates", "PX_LAST"])


def write_cpo_sheet_to_onedrive(access_token, file_path, sheet_name, df):
    print(f"\nMenyiapkan file Excel...")
    
    excel_buffer = download_excel_from_onedrive(access_token, file_path)
    
    from io import BytesIO
    output_buffer = BytesIO()
    
    if excel_buffer is None:
        print("File baru, hanya ada 1 sheet")
        with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    else:
        from openpyxl import load_workbook
        wb = load_workbook(excel_buffer)
        
        with pd.ExcelWriter(output_buffer, engine='openpyxl') as writer:
            writer.book = wb
            
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]
            
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    
    output_buffer.seek(0)
    
    print(f"Uploading ke OneDrive: {file_path}")
    upload_excel_to_onedrive(access_token, file_path, output_buffer)
    print("Upload selesai!")


def get_max_pagination(base_url):
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) "
                      "Chrome/121.0.0.0 Safari/537.36"
    }
    try:
        resp = requests.get(base_url, headers=headers, timeout=30)
        resp.raise_for_status()
    except Exception as e:
        print(f"Error: {e}")
        return 1
    
    soup = BeautifulSoup(resp.text, "lxml")
    pagination = soup.select_one("div.bdp-post-pagination")
    
    if not pagination:
        return 1
    
    page_links = pagination.select("a.page-numbers")
    max_page = 1
    
    for link in page_links:
        text = link.get_text(strip=True)
        if text.isdigit():
            max_page = max(max_page, int(text))
    
    print(f"Total halaman: {max_page}")
    return max_page


def parse_date_from_title(title):
    bulan_id = {
        "Januari": 1, "Februari": 2, "Maret": 3, "April": 4,
        "Mei": 5, "Juni": 6, "Juli": 7, "Agustus": 8,
        "September": 9, "Oktober": 10, "November": 11, "Desember": 12
    }
    
    pattern = r'(\d{1,2})\s+(Januari|Februari|Maret|April|Mei|Juni|Juli|Agustus|September|Oktober|November|Desember)\s+(\d{4})'
    match = re.search(pattern, title)
    
    if match:
        day = int(match.group(1))
        month = bulan_id[match.group(2)]
        year = int(match.group(3))
        try:
            return f"{year:04d}-{month:02d}-{day:02d}"
        except:
            return None
    return None


def scrape_articles_from_page(page_url):
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) "
                      "Chrome/121.0.0.0 Safari/537.36"
    }
    
    try:
        resp = requests.get(page_url, headers=headers, timeout=30)
        resp.raise_for_status()
    except Exception as e:
        print(f"Error di {page_url}: {e}")
        return []
    
    soup = BeautifulSoup(resp.text, "lxml")
    articles = []
    
    main_article = soup.select_one("div.bdp-left-block")
    if main_article:
        title_tag = main_article.select_one("h2.bdp-post-title a")
        if title_tag:
            title = title_tag.get_text(strip=True)
            link = title_tag.get("href", "")
            if "Posisi Harga Komoditas" in title:
                article_date = parse_date_from_title(title)
                if article_date:
                    articles.append({
                        "title": title,
                        "url": link,
                        "upload_date": article_date
                    })
    
    article_containers = soup.select("div.bdp-s-medium-9.bdp-columns")
    for container in article_containers:
        title_tag = container.select_one("h4.bdp-post-title a")
        if not title_tag:
            continue
        
        title = title_tag.get_text(strip=True)
        link = title_tag.get("href", "")
        
        if "Posisi Harga Komoditas" not in title:
            continue
        
        article_date = parse_date_from_title(title)
        if article_date:
            articles.append({
                "title": title,
                "url": link,
                "upload_date": article_date
            })
    
    return articles


def scrape_articles_until_last_date(last_date):
    base_url = "https://gapki.id/posisi-harga-komoditas/"
    print(f"\nMencari artikel baru setelah {last_date}...")
    
    max_page = get_max_pagination(base_url)
    new_articles = []
    should_stop = False
    
    for page_num in range(1, max_page + 1):
        if should_stop:
            break
        
        if page_num == 1:
            page_url = base_url
        else:
            page_url = f"{base_url}page/{page_num}/"
        
        print(f"Scraping halaman {page_num}...", end=" ")
        articles = scrape_articles_from_page(page_url)
        print(f"{len(articles)} artikel ditemukan")
        
        for article in articles:
            article_date = article["upload_date"]
            if article_date > last_date:
                new_articles.append(article)
                print(f"  {article_date} - {article['title'][:50]}...")
            else:
                print(f"  ! Artikel {article_date} sudah lebih lama dari {last_date}, berhenti")
                should_stop = True
                break
    
    print(f"\nTotal artikel baru ditemukan: {len(new_articles)}")
    return new_articles


def get_last_upload_date(access_token):
    try:
        print("\n" + "="*60)
        print("STEP 1: Membaca data existing dari OneDrive")
        print("="*60)
        print(f"File: {ONEDRIVE_FILE_PATH}")
        print(f"Sheet: {SHEET_NAME}")
        
        df = read_cpo_sheet_from_onedrive(access_token, ONEDRIVE_FILE_PATH, SHEET_NAME)
        
        if df.empty:
            print("! Sheet kosong atau tidak ada data")
            return None
        
        print(f"Berhasil membaca {len(df)} baris")
        print(f"  Kolom: {df.columns.tolist()}")
        
        if "Upload_Dates" not in df.columns:
            print("Kolom 'Upload_Dates' tidak ditemukan!")
            print(f"  Kolom yang tersedia: {df.columns.tolist()}")
            return None
        
        df["Upload_Dates"] = pd.to_datetime(df["Upload_Dates"], errors='coerce')
        df_valid = df.dropna(subset=["Upload_Dates"])
        
        if df_valid.empty:
            print("! Tidak ada tanggal valid di kolom Upload_Dates")
            return None
        
        last_date = df_valid["Upload_Dates"].max()
        last_date_str = last_date.strftime("%Y-%m-%d")
        print(f"Upload_Dates terakhir: {last_date_str}")
        print(f"  Total baris dengan tanggal valid: {len(df_valid)}")
        
        return last_date_str
        
    except Exception as e:
        print(f"Error membaca sheet: {type(e).__name__}")
        print(f"  Message: {str(e)}")
        import traceback
        traceback.print_exc()
        return None


def parse_date_in_parentheses(date_str, full_text, article_title=None):
    bulan_abbr = {
        "Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4,
        "Mei": 5, "Jun": 6, "Jul": 7, "Agt": 8, "Agu": 8,
        "Sep": 9, "Okt": 10, "Nov": 11, "Des": 12
    }
    
    match = re.search(r"(\d{1,2})\s*(\w+)", date_str)
    if not match:
        return None
    
    day = int(match.group(1))
    month_str = match.group(2)
    month = None
    
    for abbr, num in bulan_abbr.items():
        if month_str.startswith(abbr):
            month = num
            break
    
    if not month:
        return None
    
    year = None
    if article_title:
        year_match = re.search(r'20\d{2}', article_title)
        if year_match:
            year = int(year_match.group(0))
    
    if not year:
        year_match = re.search(r'20\d{2}', full_text)
        if year_match:
            year = int(year_match.group(0))
    
    if not year:
        year = datetime.now().year
    
    try:
        return f"{year:04d}-{month:02d}-{day:02d}"
    except:
        return None


def scrape_harga_multi(url, article_title=None):
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) "
                      "Chrome/121.0.0.0 Safari/537.36"
    }
    
    try:
        resp = requests.get(url, headers=headers, timeout=15)
        resp.raise_for_status()
    except Exception as e:
        print(f"  Error mengakses artikel: {e}")
        return []
    
    soup = BeautifulSoup(resp.text, "lxml")
    paragraphs = soup.select("div.nv-content-wrap.entry-content p")
    harga_list = []
    
    for p in paragraphs:
        text = p.get_text()
        lines = text.split('\n')
        
        for line in lines:
            line = line.strip()
            
            if "KPB" in line.upper() and "CPO" in line.upper():
                pattern_with_date = r"([\d\.]+)\s*\((\d{1,2})\s*(\w+)['\u2019]?\)"
                matches = list(re.finditer(pattern_with_date, line))
                
                if matches:
                    for match in matches:
                        val = re.sub(r"[^\d]", "", match.group(1))
                        try:
                            harga = int(val)
                            day = match.group(2)
                            month_abbr = match.group(3)
                            date_str = f"{day} {month_abbr}"
                            parsed_date = parse_date_in_parentheses(date_str, line, article_title)
                            
                            harga_list.append({
                                "harga": harga,
                                "date_str": date_str,
                                "parsed_date": parsed_date
                            })
                            print(f"    Harga: {harga} | Tanggal: ({date_str}) -> {parsed_date}")
                        except Exception as e:
                            continue
                    
                    return harga_list
                else:
                    match = re.search(r"IDR\s*([\d\.,]+)", line)
                    if match:
                        val = re.sub(r"[^\d]", "", match.group(1))
                        try:
                            harga = int(val)
                            harga_list.append({
                                "harga": harga,
                                "date_str": None,
                                "parsed_date": None
                            })
                            print(f"    Harga tanpa tanggal: {harga}")
                            return harga_list
                        except:
                            continue
    
    if not harga_list:
        print("    ! Tidak ditemukan harga")
    
    return harga_list


def update_onedrive_with_new_data(access_token, new_data_list):
    print("\n" + "="*60)
    print("STEP 3: Menyimpan data ke OneDrive")
    print("="*60)
    
    df_old = read_cpo_sheet_from_onedrive(access_token, ONEDRIVE_FILE_PATH, SHEET_NAME)
    
    if not df_old.empty:
        print(f"Data lama: {len(df_old)} baris")
    else:
        df_old = pd.DataFrame(columns=["Upload_Dates", "Dates", "PX_LAST"])
        print("! Sheet kosong, akan membuat baru")
    
    df_new = pd.DataFrame(new_data_list)
    print(f"Data baru: {len(df_new)} baris")
    
    df_final = pd.concat([df_old, df_new], ignore_index=True)
    
    df_final["Upload_Dates"] = pd.to_datetime(df_final["Upload_Dates"], errors='coerce')
    df_final["Dates"] = pd.to_datetime(df_final["Dates"], errors='coerce')
    df_final.drop_duplicates(subset=["Dates"], keep="last", inplace=True)
    
    df_final["Upload_Dates"] = df_final["Upload_Dates"].dt.date
    df_final["Dates"] = df_final["Dates"].dt.date
    
    df_final.sort_values("Dates", ascending=True, inplace=True)
    
    print(f"Data setelah deduplikasi: {len(df_final)} baris")
    
    write_cpo_sheet_to_onedrive(access_token, ONEDRIVE_FILE_PATH, SHEET_NAME, df_final)
    
    print(f"\n{'='*60}")
    print("DATA BERHASIL DISIMPAN KE ONEDRIVE")
    print(f"{'='*60}")
    print(f"  File: {ONEDRIVE_FILE_PATH}")
    print(f"  Sheet: {SHEET_NAME}")
    print(f"  Total rows: {len(df_final)}")
    print(f"  Data baru ditambahkan: {len(new_data_list)} baris")


def main_scraper_cpo():
    print("\n" + "="*60)
    print("GAPKI CPO PRICE SCRAPER")
    print("STORAGE MODE: OneDrive")
    print("="*60)
    
    print(f"\nFile: {ONEDRIVE_FILE_PATH}")
    print(f"Sheet: {SHEET_NAME}")
    
    print("\nAuthenticating to Microsoft Graph API...")
    try:
        access_token = get_access_token()
        print("Authentication successful")
    except Exception as e:
        print(f"Authentication failed: {e}")
        return
    
    last_date = get_last_upload_date(access_token)
    if not last_date:
        print("\nTidak bisa melanjutkan tanpa tanggal terakhir")
        print("  Pastikan sheet sudah ada dan memiliki data dengan kolom 'Upload_Dates'")
        return
    
    print("\n" + "="*60)
    print("STEP 2: Scraping artikel baru")
    print("="*60)
    
    new_articles = scrape_articles_until_last_date(last_date)
    if not new_articles:
        print("\n! Tidak ada artikel baru untuk di-scrape")
        return
    
    print(f"\n{'='*60}")
    print(f"Mengambil harga dari {len(new_articles)} artikel...")
    print(f"{'='*60}")
    
    all_data = []
    
    for idx, article in enumerate(new_articles, 1):
        print(f"\n[{idx}/{len(new_articles)}] {article['title'][:60]}...")
        print(f"  URL: {article['url']}")
        harga_list = scrape_harga_multi(article['url'], article['title'])
        
        if not harga_list:
            print(f"  ! Skip - tidak ada harga ditemukan")
            continue
        
        if len(harga_list) > 1:
            print(f"  Ditemukan {len(harga_list)} harga berbeda")
        
        for harga_data in harga_list:
            if harga_data["parsed_date"]:
                dates = harga_data["parsed_date"]
            else:
                dates = article["upload_date"]
            
            all_data.append({
                "Upload_Dates": article["upload_date"],
                "Dates": dates,
                "PX_LAST": harga_data["harga"]
            })
    
    if all_data:
        update_onedrive_with_new_data(access_token, all_data)
        print("\n" + "="*60)
        print("SELESAI! Data berhasil diupdate")
        print("="*60)
    else:
        print("\n! Tidak ada data untuk disimpan")


if __name__ == "__main__":
    main_scraper_cpo()