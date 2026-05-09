import os
import re
import sys
from datetime import datetime
from io import BytesIO

import pandas as pd
import requests
from bs4 import BeautifulSoup
from dotenv import load_dotenv
from openpyxl import load_workbook

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from helpers.onedrive_helper import (
    download_excel_from_onedrive,
    get_access_token,
    upload_excel_to_onedrive,
)

load_dotenv()


# Constants

ONEDRIVE_FILE_PATH = os.getenv("ONEDRIVE_DATA_PATH", "/results/(Terstruktur)Data_Scraping_final.xlsx")
SHEET_NAME         = "(Data)CPO"

GAPKI_BASE_URL = "https://gapki.id/posisi-harga-komoditas/"

REQUEST_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/121.0.0.0 Safari/537.36"
    )
}
REQUEST_TIMEOUT = 30

MONTHS_ID_ABBR = {
    "Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4,
    "Mei": 5, "Jun": 6, "Jul": 7, "Agt": 8, "Agu": 8,
    "Sep": 9, "Okt": 10, "Nov": 11, "Des": 12,
}
MONTHS_ID_FULL = {
    "Januari": 1,  "Februari": 2,  "Maret": 3,    "April": 4,
    "Mei": 5,      "Juni": 6,      "Juli": 7,      "Agustus": 8,
    "September": 9,"Oktober": 10,  "November": 11, "Desember": 12,
}


# Date Utilities

def parse_date_from_title(title):
    """
    Parse full Indonesian date from article title.
    e.g. 'Posisi Harga Komoditas 5 Januari 2025' → '2025-01-05'
    """
    pattern = (
        r"(\d{1,2})\s+"
        r"(Januari|Februari|Maret|April|Mei|Juni|Juli|Agustus"
        r"|September|Oktober|November|Desember)\s+(\d{4})"
    )
    match = re.search(pattern, title)
    if match:
        day   = int(match.group(1))
        month = MONTHS_ID_FULL[match.group(2)]
        year  = int(match.group(3))
        try:
            return f"{year:04d}-{month:02d}-{day:02d}"
        except Exception:
            return None
    return None

def parse_date_in_parentheses(date_str, full_text, article_title=None):
    """
    Parse abbreviated Indonesian date from parenthesized price notation.
    e.g. '(5 Jan)' within a price line → '2025-01-05'
    """
    match = re.search(r"(\d{1,2})\s*(\w+)", date_str)
    if not match:
        return None

    day       = int(match.group(1))
    month_str = match.group(2)
    month     = None

    for abbr, num in MONTHS_ID_ABBR.items():
        if month_str.startswith(abbr):
            month = num
            break

    if not month:
        return None

    year = None
    if article_title:
        year_match = re.search(r"20\d{2}", article_title)
        if year_match:
            year = int(year_match.group(0))

    if not year:
        year_match = re.search(r"20\d{2}", full_text)
        if year_match:
            year = int(year_match.group(0))

    if not year:
        year = datetime.now().year

    try:
        return f"{year:04d}-{month:02d}-{day:02d}"
    except Exception:
        return None


# OneDrive Read / Write

def read_cpo_sheet_from_onedrive(access_token):
    """
    Download and return the CPO sheet from OneDrive as a DataFrame.

    Returns an empty DataFrame with correct columns if file or sheet is missing.
    """
    excel_buffer = download_excel_from_onedrive(access_token, ONEDRIVE_FILE_PATH)

    if excel_buffer is None:
        print("[Read] File tidak ditemukan, akan membuat baru.")
        return pd.DataFrame(columns=["Upload_Dates", "Dates", "PX_LAST"])

    try:
        df = pd.read_excel(excel_buffer, sheet_name=SHEET_NAME)
        print(f"[Read] Berhasil baca sheet '{SHEET_NAME}', rows={len(df)}.")
        return df
    except Exception as exc:
        print(f"[Read] Sheet '{SHEET_NAME}' tidak ditemukan: {exc} — akan membuat baru.")
        return pd.DataFrame(columns=["Upload_Dates", "Dates", "PX_LAST"])

def write_cpo_sheet_to_onedrive(access_token, df):
    """
    Upload the CPO DataFrame to OneDrive, preserving all other sheets.
    """
    print(f"\n[Write] Menyiapkan file Excel...")
    excel_buffer  = download_excel_from_onedrive(access_token, ONEDRIVE_FILE_PATH)
    output_buffer = BytesIO()

    try:
        if excel_buffer is None:
            print("[Write] File baru — hanya ada 1 sheet.")
            with pd.ExcelWriter(output_buffer, engine="openpyxl", mode="w") as writer:
                df.to_excel(writer, sheet_name=SHEET_NAME, index=False)
        else:
            print("[Write] File existing — preserve semua sheet...")
            excel_buffer.seek(0)
            wb = load_workbook(excel_buffer)
            print(f"[Write] Sheet saat ini: {wb.sheetnames}")

            # Fix hidden sheets
            visible_sheets = [s for s in wb.worksheets if s.sheet_state == "visible"]
            if len(visible_sheets) == 0:
                print("[Write] Fixing hidden sheets...")
                wb.worksheets[0].sheet_state = "visible"
                wb.active = 0
            for sheet in wb.worksheets:
                sheet.sheet_state = "visible"

            # Hapus dan buat ulang sheet target
            if SHEET_NAME in wb.sheetnames:
                print(f"[Write] Menghapus sheet '{SHEET_NAME}' yang lama...")
                del wb[SHEET_NAME]
            ws = wb.create_sheet(SHEET_NAME)

            for col_idx, col_name in enumerate(df.columns, 1):
                ws.cell(row=1, column=col_idx, value=col_name)
            for row_idx, row_data in enumerate(df.values, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)

            print(f"[Write] Sheet yang akan disimpan: {wb.sheetnames}")
            wb.save(output_buffer)
            wb.close()

        output_buffer.seek(0)

        # Verifikasi
        verify_wb = load_workbook(output_buffer)
        print(f"[Write] Verifikasi sheet: {verify_wb.sheetnames}")
        verify_wb.close()
        output_buffer.seek(0)

        print(f"[Write] Uploading ke OneDrive: {ONEDRIVE_FILE_PATH}")
        upload_excel_to_onedrive(access_token, ONEDRIVE_FILE_PATH, output_buffer)
        print("[Write] Upload selesai!")

    except Exception as exc:
        print(f"[Write] Error saat menyimpan: {exc}")
        import traceback
        traceback.print_exc()
        raise


# Last Upload Date Check

def get_last_upload_date(access_token):
    """
    Read the CPO sheet and return the latest Upload_Dates as a YYYY-MM-DD string.

    Returns None if the sheet is empty or unreadable.
    """
    print(f"\n{'='*60}")
    print("[Check] Membaca data existing dari OneDrive")
    print(f"{'='*60}")
    print(f"[Check] File : {ONEDRIVE_FILE_PATH}")
    print(f"[Check] Sheet: {SHEET_NAME}")

    df = read_cpo_sheet_from_onedrive(access_token)

    if df.empty:
        print("[Check] Sheet kosong atau tidak ada data.")
        return None

    if "Upload_Dates" not in df.columns:
        print(f"[Check] Kolom 'Upload_Dates' tidak ditemukan. Kolom: {df.columns.tolist()}")
        return None

    df["Upload_Dates"] = pd.to_datetime(df["Upload_Dates"], errors="coerce")
    df_valid = df.dropna(subset=["Upload_Dates"])

    if df_valid.empty:
        print("[Check] Tidak ada tanggal valid di kolom Upload_Dates.")
        return None

    last_date     = df_valid["Upload_Dates"].max()
    last_date_str = last_date.strftime("%Y-%m-%d")
    print(f"[Check] Upload_Dates terakhir : {last_date_str}")
    print(f"[Check] Total baris valid     : {len(df_valid)}")
    return last_date_str


# Article Scraping

def _get_max_pagination(base_url):
    """Return the total number of pagination pages on the GAPKI listing page."""
    try:
        resp = requests.get(base_url, headers=REQUEST_HEADERS, timeout=REQUEST_TIMEOUT)
        resp.raise_for_status()
    except Exception as exc:
        print(f"[Pagination] Error: {exc}")
        return 1

    soup       = BeautifulSoup(resp.text, "lxml")
    pagination = soup.select_one("div.bdp-post-pagination")

    if not pagination:
        return 1

    max_page = 1
    for link in pagination.select("a.page-numbers"):
        text = link.get_text(strip=True)
        if text.isdigit():
            max_page = max(max_page, int(text))

    print(f"[Pagination] Total halaman: {max_page}")
    return max_page

def _scrape_articles_from_page(page_url):
    """
    Scrape article titles and URLs from a single GAPKI listing page.

    Returns a list of dicts with title, url, and upload_date.
    """
    try:
        resp = requests.get(page_url, headers=REQUEST_HEADERS, timeout=REQUEST_TIMEOUT)
        resp.raise_for_status()
    except Exception as exc:
        print(f"[Scrape] Error di {page_url}: {exc}")
        return []

    soup     = BeautifulSoup(resp.text, "lxml")
    articles = []

    # Main featured article
    main_article = soup.select_one("div.bdp-left-block")
    if main_article:
        title_tag = main_article.select_one("h2.bdp-post-title a")
        if title_tag:
            title = title_tag.get_text(strip=True)
            link  = title_tag.get("href", "")
            if "Posisi Harga Komoditas" in title:
                article_date = parse_date_from_title(title)
                if article_date:
                    articles.append({"title": title, "url": link, "upload_date": article_date})

    # Secondary articles
    for container in soup.select("div.bdp-s-medium-9.bdp-columns"):
        title_tag = container.select_one("h4.bdp-post-title a")
        if not title_tag:
            continue
        title = title_tag.get_text(strip=True)
        link  = title_tag.get("href", "")
        if "Posisi Harga Komoditas" not in title:
            continue
        article_date = parse_date_from_title(title)
        if article_date:
            articles.append({"title": title, "url": link, "upload_date": article_date})

    return articles

def scrape_articles_until_last_date(last_date):
    """
    Crawl GAPKI listing pages and collect articles newer than last_date.

    Stops early once an article dated <= last_date is encountered.
    """
    print(f"\n[Scrape] Mencari artikel baru setelah {last_date}...")

    max_page     = _get_max_pagination(GAPKI_BASE_URL)
    new_articles = []
    should_stop  = False

    for page_num in range(1, max_page + 1):
        if should_stop:
            break

        page_url = GAPKI_BASE_URL if page_num == 1 else f"{GAPKI_BASE_URL}page/{page_num}/"
        print(f"[Scrape] Halaman {page_num}...", end=" ")
        articles = _scrape_articles_from_page(page_url)
        print(f"{len(articles)} artikel ditemukan.")

        for article in articles:
            article_date = article["upload_date"]
            if article_date > last_date:
                new_articles.append(article)
                print(f"[Scrape] + {article_date} — {article['title'][:50]}...")
            else:
                print(f"[Scrape] ! {article_date} sudah lama dari {last_date} — berhenti.")
                should_stop = True
                break

    print(f"\n[Scrape] Total artikel baru: {len(new_articles)}")
    return new_articles


# Price Scraping

def scrape_harga_multi(url, article_title=None):
    """
    Fetch and parse CPO prices from a single GAPKI article page.

    Returns a list of dicts with harga, date_str, and parsed_date.
    """
    try:
        resp = requests.get(url, headers=REQUEST_HEADERS, timeout=15)
        resp.raise_for_status()
    except Exception as exc:
        print(f"[Price] Error mengakses artikel: {exc}")
        return []

    soup       = BeautifulSoup(resp.text, "lxml")
    paragraphs = soup.select("div.nv-content-wrap.entry-content p")
    harga_list = []

    for p in paragraphs:
        text  = p.get_text()
        lines = text.split("\n")

        for line in lines:
            line = line.strip()

            if "KPB" not in line.upper() or "CPO" not in line.upper():
                continue

            # Pattern: nilai (tanggal) e.g. "14.958 (5 Jan)"
            matches = list(re.finditer(
                r"([\d\.]+)\s*\((\d{1,2})\s*(\w+)['\u2019]?\)",
                line,
            ))

            if matches:
                for match in matches:
                    val = re.sub(r"[^\d]", "", match.group(1))
                    try:
                        harga      = int(val)
                        day        = match.group(2)
                        month_abbr = match.group(3)
                        date_str   = f"{day} {month_abbr}"
                        parsed_date = parse_date_in_parentheses(date_str, line, article_title)
                        harga_list.append({
                            "harga":       harga,
                            "date_str":    date_str,
                            "parsed_date": parsed_date,
                        })
                        print(f"[Price] Harga: {harga} | ({date_str}) → {parsed_date}")
                    except Exception:
                        continue
                return harga_list

            # Fallback: IDR pattern e.g. "IDR 14.958"
            match = re.search(r"IDR\s*([\d\.,]+)", line)
            if match:
                val = re.sub(r"[^\d]", "", match.group(1))
                try:
                    harga = int(val)
                    harga_list.append({"harga": harga, "date_str": None, "parsed_date": None})
                    print(f"[Price] Harga tanpa tanggal: {harga}")
                    return harga_list
                except Exception:
                    continue

    if not harga_list:
        print("[Price] Tidak ditemukan harga.")

    return harga_list


# Save to OneDrive

def update_onedrive_with_new_data(access_token, new_data_list):
    """
    Merge new CPO price data with existing OneDrive sheet, deduplicate, sort, and upload.
    """
    print(f"\n{'='*60}")
    print("[Save] Menyimpan data ke OneDrive")
    print(f"{'='*60}")

    df_old = read_cpo_sheet_from_onedrive(access_token)

    if not df_old.empty:
        print(f"[Save] Data lama: {len(df_old)} baris.")
    else:
        df_old = pd.DataFrame(columns=["Upload_Dates", "Dates", "PX_LAST"])
        print("[Save] Sheet kosong — akan membuat baru.")

    df_new   = pd.DataFrame(new_data_list)
    print(f"[Save] Data baru: {len(df_new)} baris.")

    df_final = pd.concat([df_old, df_new], ignore_index=True)

    df_final["Upload_Dates"] = pd.to_datetime(df_final["Upload_Dates"], errors="coerce")
    df_final["Dates"]        = pd.to_datetime(df_final["Dates"], errors="coerce")
    df_final.drop_duplicates(subset=["Dates"], keep="last", inplace=True)
    df_final["Upload_Dates"] = df_final["Upload_Dates"].dt.date
    df_final["Dates"]        = df_final["Dates"].dt.date
    df_final.sort_values("Dates", ascending=True, inplace=True)

    print(f"[Save] Data setelah deduplikasi: {len(df_final)} baris.")

    write_cpo_sheet_to_onedrive(access_token, df_final)

    print(f"\n{'='*60}")
    print("[Save] DATA BERHASIL DISIMPAN KE ONEDRIVE")
    print(f"{'='*60}")
    print(f"[Save] File      : {ONEDRIVE_FILE_PATH}")
    print(f"[Save] Sheet     : {SHEET_NAME}")
    print(f"[Save] Total rows: {len(df_final)}")
    print(f"[Save] Data baru : {len(new_data_list)} baris")


# Public Entry Point

def main_scraper_cpo():
    """
    Run the full CPO price scraping workflow:
    authenticate, check last date, scrape new articles, extract prices, save to OneDrive.
    """
    print(f"\n{'='*60}")
    print("GAPKI CPO PRICE SCRAPER")
    print("STORAGE MODE: OneDrive")
    print(f"{'='*60}")
    print(f"\n[Main] File : {ONEDRIVE_FILE_PATH}")
    print(f"[Main] Sheet: {SHEET_NAME}")

    print("\n[Main] Authenticating to Microsoft Graph API...")
    try:
        access_token = get_access_token()
        print("[Main] Authentication successful.")
    except Exception as exc:
        print(f"[Main] Authentication failed: {exc}")
        return

    last_date = get_last_upload_date(access_token)
    if not last_date:
        print("\n[Main] Sheet kosong — scrape semua artikel yang tersedia.")
        last_date = "2000-01-01"

        print(f"\n{'='*60}")
        print("[Main] Scraping artikel baru")
        print(f"{'='*60}")

    new_articles = scrape_articles_until_last_date(last_date)
    if not new_articles:
        print("\n[Main] Tidak ada artikel baru untuk di-scrape.")
        return

    print(f"\n{'='*60}")
    print(f"[Main] Mengambil harga dari {len(new_articles)} artikel...")
    print(f"{'='*60}")

    all_data = []

    for idx, article in enumerate(new_articles, 1):
        print(f"\n[Main] ({idx}/{len(new_articles)}) {article['title'][:60]}...")
        print(f"[Main] URL: {article['url']}")
        harga_list = scrape_harga_multi(article["url"], article["title"])

        if not harga_list:
            print("[Main] Skip — tidak ada harga ditemukan.")
            continue

        if len(harga_list) > 1:
            print(f"[Main] Ditemukan {len(harga_list)} harga berbeda.")

        for harga_data in harga_list:
            dates = harga_data["parsed_date"] or article["upload_date"]
            all_data.append({
                "Upload_Dates": article["upload_date"],
                "Dates":        dates,
                "PX_LAST":      harga_data["harga"],
            })

    if all_data:
        update_onedrive_with_new_data(access_token, all_data)
        print(f"\n{'='*60}")
        print("[Main] SELESAI!")
        print(f"{'='*60}\n")
    else:
        print("\n[Main] Tidak ada data untuk disimpan.")


# Script Entry Point

if __name__ == "__main__":
    main_scraper_cpo()