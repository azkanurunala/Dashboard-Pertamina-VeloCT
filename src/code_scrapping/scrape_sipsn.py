import os
import sys
import requests
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from dotenv import load_dotenv
import traceback
from datetime import datetime

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from helpers.onedrive_helper import (
    get_access_token,
    download_excel_from_onedrive,
    upload_excel_to_onedrive
)

load_dotenv()

ONEDRIVE_FILE_PATH = os.getenv("ONEDRIVE_DATA_PATH", "/results/(Terstruktur)Data Scrapping.xlsx")
API_URL = "https://sipsn.kemenlh.go.id/sipsn/public/home/ajax_list"
SHEET_MAPPING = {
    'sumber': '(Data)WTE_Sumber',
    'komposisi': '(Data)WTE_Komposisi',
    'timbulan': '(Data)WTE_Timbulan'
}

def clean_column_names(df):
    rename_dict = {
        'nama_dati2': 'Nama Kota/Kabupaten',
        'nama_propinsi': 'Nama Provinsi'
    }
    df.rename(columns=rename_dict, inplace=True)
    return df

def fetch_all_data(tahun: str = '2025'):
    jenis_list = ['sumber', 'komposisi', 'timbulan']
    all_data = {}
    for jenis in jenis_list:
        print(f"Mengambil data {jenis} tahun {tahun}...")
        payload = {
            'length': '-1',
            'jenis': jenis,
            'tahun': tahun,
        }
        response = requests.post(API_URL, data=payload)
        if response.status_code == 200:
            data = response.json()
            df = pd.DataFrame(data['data'])
            df = clean_column_names(df)
            all_data[jenis] = df
            print(f"Data {jenis}: {len(df)} records")
        else:
            print(f"Gagal ambil data {jenis}: {response.status_code}")
    return all_data

def merge_data_by_year(df_existing, df_new):
    df_combined = pd.concat([df_existing, df_new], ignore_index=True)
    id_cols = []
    if 'tahun' in df_combined.columns:
        id_cols.append('tahun')
    if 'Nama Provinsi' in df_combined.columns:
        id_cols.append('Nama Provinsi')
    if 'Nama Kota/Kabupaten' in df_combined.columns:
        id_cols.append('Nama Kota/Kabupaten')
    if id_cols:
        df_combined = df_combined.drop_duplicates(subset=id_cols, keep='last')
    if 'tahun' in df_combined.columns:
        df_combined['tahun'] = pd.to_numeric(df_combined['tahun'], errors='coerce')
        df_combined = df_combined.sort_values('tahun', ascending=True)
    print(f"Data setelah merge: {len(df_combined)} rows")
    return df_combined

def save_to_onedrive(access_token, data_dict: dict, tahun: str):
    print(f"\n{'='*60}")
    print(f"Menyimpan data tahun {tahun} ke OneDrive...")
    print(f"{'='*60}")
    if not data_dict or all(df.empty for df in data_dict.values()):
        print("Tidak ada data untuk disimpan")
        return
    excel_buffer = download_excel_from_onedrive(access_token, ONEDRIVE_FILE_PATH)
    output_buffer = BytesIO()
    try:
        if excel_buffer is None:
            with pd.ExcelWriter(output_buffer, engine='openpyxl', mode='w') as writer:
                for jenis, df in data_dict.items():
                    sheet_name = SHEET_MAPPING[jenis]
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
        else:
            print("File ditemukan di OneDrive, merging data...\n")
            excel_buffer.seek(0)
            wb = load_workbook(excel_buffer)
            visible_sheets = [s for s in wb.worksheets if s.sheet_state == 'visible']
            if len(visible_sheets) == 0:
                wb.worksheets[0].sheet_state = 'visible'
                wb.active = 0
            for jenis, df_new in data_dict.items():
                sheet_name = SHEET_MAPPING[jenis]
                print(f"[{sheet_name}]")
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    first_col = 1
                    while first_col <= ws.max_column and ws.cell(1, first_col).value is None:
                        first_col += 1
                    headers = []
                    for col in range(first_col, ws.max_column + 1):
                        headers.append(ws.cell(1, col).value)
                    existing_data = []
                    for row in range(2, ws.max_row + 1):
                        row_data = []
                        for col in range(first_col, ws.max_column + 1):
                            row_data.append(ws.cell(row, col).value)
                        existing_data.append(row_data)
                    df_existing = pd.DataFrame(existing_data, columns=headers)
                    print(f"  Data existing: {len(df_existing)} rows")
                    print(f"  Data baru (tahun {tahun}): {len(df_new)} rows")
                    df_merged = merge_data_by_year(df_existing, df_new)
                    del wb[sheet_name]
                    ws = wb.create_sheet(sheet_name)
                    for col_idx, col_name in enumerate(df_merged.columns, 1):
                        ws.cell(row=1, column=col_idx, value=col_name)
                    for row_idx, row_data in enumerate(df_merged.values, 2):
                        for col_idx, value in enumerate(row_data, 1):
                            ws.cell(row=row_idx, column=col_idx, value=value)
                    print(f"Merged: total {len(df_merged)} rows\n")
                else:
                    print(f"Sheet belum ada, membuat baru...")
                    ws = wb.create_sheet(sheet_name)
                    for col_idx, col_name in enumerate(df_new.columns, 1):
                        ws.cell(row=1, column=col_idx, value=col_name)
                    for row_idx, row_data in enumerate(df_new.values, 2):
                        for col_idx, value in enumerate(row_data, 1):
                            ws.cell(row=row_idx, column=col_idx, value=value)
                    print(f"Created: {len(df_new)} rows\n")
            wb.save(output_buffer)
            wb.close()
        output_buffer.seek(0)
        verify_wb = load_workbook(output_buffer)
        for sheet_name in verify_wb.sheetnames:
            ws = verify_wb[sheet_name]
        verify_wb.close()
        output_buffer.seek(0)
        print(f"\n{'='*60}")
        print(f"Uploading ke OneDrive: {ONEDRIVE_FILE_PATH}")
        print(f"{'='*60}")
        upload_excel_to_onedrive(access_token, ONEDRIVE_FILE_PATH, output_buffer)
    except Exception as e:
        traceback.print_exc()

def main_sipsn_scraper():
    try:
        access_token = get_access_token()
        print("Authentication successful")
    except Exception as e:
        print(f"Authentication failed: {e}")
        return
    tahun = str(datetime.now().year-1)
    data_dict = fetch_all_data(tahun=tahun)
    if data_dict:
        print(f"\nSaving to OneDrive...")
        save_to_onedrive(access_token, data_dict, tahun)
    else:
        print("\nTidak ada data yang berhasil diambil")

if __name__ == "__main__":
    main_sipsn_scraper()